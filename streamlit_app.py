from __future__ import annotations

# Limit native math runtimes before NumPy/XGBoost are imported. Streamlit Cloud
# otherwise allows several OpenMP/BLAS pools to compete in one small process,
# which can terminate the interpreter with a segmentation fault.
import os
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")
os.environ.setdefault("NUMEXPR_NUM_THREADS", "1")

import traceback
import warnings
from typing import Any

import logging
import sys

logger = logging.getLogger(__name__)
from pathlib import Path

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import pandas as pd
import streamlit as st

from core.streamlit_pipeline import (
    generate_parlays,
    optimize_portfolio_allocation,
    run_analysis_pipeline,
    run_bankroll_simulation,
    PIPELINE_BUILD,
    CANONICAL_BET_COLUMNS,
    VALID_MARKETS,
    MIN_EDGE_THRESHOLD,
    ensure_best_pick_export_columns,
    REQUIRED_BEST_PICK_EXPORT_COLUMNS,
    apply_no_bet_pick_quality,
)
from core.team_normalizer import normalize_team
from core.theover_loader import load_theover_csv
from app_core.weights_config import (
    ENABLE_EMPTY_CARD_RECOVERY,
    EMPTY_CARD_RECOVERY_MAX_PICKS,
    EMPTY_CARD_RECOVERY_MIN_PRODUCTION_EV,
    EMPTY_CARD_RECOVERY_MIN_PRODUCTION_EDGE,
    EMPTY_CARD_RECOVERY_MIN_PRODUCTION_WIN_PROB,
    EMPTY_CARD_RECOVERY_EXCLUDE_MARKET_TYPES,
    EMPTY_CARD_RECOVERY_EXCLUDE_SOURCES,
    EMPTY_CARD_RECOVERY_MAX_KELLY_TOTAL_PCT,
    EMPTY_CARD_RECOVERY_MAX_KELLY_PER_PICK_PCT,
    ALLOW_MLB_TOTAL_OVER_EMPTY_CARD_RECOVERY,
)

try:
    from app_core.kalshi_integrator import enrich_with_kalshi_markets as _enrich_kalshi_raw
except Exception:  # pragma: no cover
    _enrich_kalshi_raw = None  # type: ignore[assignment]

try:
    from app_core.odds_api import OddsAPIAuthError
except ImportError:
    class OddsAPIAuthError(Exception): pass


KALSHI_ENRICH_TIMEOUT_SECONDS = 450


def _enrich_with_kalshi_safe(df: pd.DataFrame) -> tuple[pd.DataFrame, str | None]:
    """Run Kalshi enrichment with a hard timeout.
    Returns (enriched_df, error_message_or_None).
    """
    if _enrich_kalshi_raw is None:
        return df, None

    import concurrent.futures

    executor = concurrent.futures.ThreadPoolExecutor(max_workers=1)
    future = executor.submit(_enrich_kalshi_raw, df)
    try:
        result = future.result(timeout=KALSHI_ENRICH_TIMEOUT_SECONDS)
        return result, None
    except concurrent.futures.TimeoutError:
        future.cancel()
        return df, f"Kalshi enrichment timed out (>{KALSHI_ENRICH_TIMEOUT_SECONDS}s) — skipped."
    except Exception as e:
        failing_game_ids: list[str] = []
        try:
            if isinstance(df, pd.DataFrame) and not df.empty:
                if "game_id" in df.columns:
                    failing_game_ids = (
                        df["game_id"].dropna().astype(str).head(5).tolist()
                    )
                elif {"league", "home_team", "away_team", "game_date"}.issubset(df.columns):
                    preview = (
                        df[["league", "home_team", "away_team", "game_date"]]
                        .head(3)
                        .to_dict(orient="records")
                    )
                    logger.error("Kalshi enrichment failed on rows preview: %s", preview)
        except Exception:
            failing_game_ids = []

        logger.error(
            "Kalshi enrichment failed: %s | sample_game_ids=%s",
            e,
            failing_game_ids,
            exc_info=True,
        )
        return df, f"Kalshi enrichment failed: {e}\n{traceback.format_exc()}"
    finally:
        executor.shutdown(wait=False, cancel_futures=True)


def _safe_str_series(df: pd.DataFrame, col: str, default: str = "") -> pd.Series:
    if df is None:
        return pd.Series(dtype="string")
    if df.empty:
        return pd.Series(index=df.index, dtype="string")
    if col in df.columns:
        series = df[col]
        if isinstance(series.dtype, pd.CategoricalDtype):
            series = series.astype("object")
        return series.astype("string").fillna(default)
    return pd.Series([default] * len(df), index=df.index, dtype="string")






_NO_BET_STAGE_REASONS = {
    "game_already_started": "⏱️ Game already started — live odds, not pre-game lines",
    "baseline_guardrail": "Market priced better — negative EV at this price",
    "actionable_threshold": "Failed the confidence threshold",
    "min_win_probability_floor": "Below the 55% win-probability floor",
    "empirical_proven_losing_bucket": "Proven-losing bucket — history says pass",
    "empirical_tier_overlay": "Small-sample bucket — not enough history to trust",
    "divergence_viability_floor": "Signals diverge and the pick failed the viability floor",
    "kalshi_wrong_game_title": "Kalshi data mismatch (wrong game)",
    "line_provenance_unresolved": "Could not verify the betting line",
    "extreme_price_guard": "Odds moved to an extreme price",
}


# User-facing status vocabulary (owner, 4 Jul): "No Play" reads like the game is
# off the board when it means "no edge at this price", and most of the card
# carried it. Everything the user READS gets these labels; the internal gate
# names ("No Play"/"Below Threshold"/...) stay untouched inside the pipeline,
# where grading history and gate logic key off them.
STATUS_DISPLAY_LABELS = {
    "No Play": "No Edge",
    "Below Threshold": "Near Miss",
    "High Variance/Speculative": "Unproven",
    "Fallback / Low Confidence": "Low Data",
}


def apply_status_display_labels(df: pd.DataFrame) -> pd.DataFrame:
    """Map internal status names to the user-facing vocabulary (display/export only)."""
    if df is None or df.empty:
        return df
    out = df.copy()
    if "Pick_Status" in out.columns:
        out["Pick_Status"] = out["Pick_Status"].astype(str).replace(STATUS_DISPLAY_LABELS)
    for col in ("Pick_Quality", "Pick Quality"):
        if col in out.columns:
            s = out[col].astype(str)
            for old, new in STATUS_DISPLAY_LABELS.items():
                s = s.str.replace(old, new, regex=False)
            out[col] = s
    return out


def _friendly_no_bet_reason(row: pd.Series | dict) -> str:
    """One plain-English line for why a non-Actionable pick isn't bettable.

    Display-only: collapses the status ladder (No Play / Below Threshold /
    High Variance) into a single 'why', keyed on status_blocker_stage with the
    raw Status_Reason as fallback. Exports and grading keep the raw statuses.
    """
    get = row.get if hasattr(row, "get") else lambda k, d="": d
    stage = str(get("status_blocker_stage", "") or "").strip()
    if stage in _NO_BET_STAGE_REASONS:
        return _NO_BET_STAGE_REASONS[stage]
    reason = str(get("Status_Reason", "") or get("status_blocker_reason", "") or "").strip()
    return reason if reason and reason.lower() != "nan" else "Did not qualify"


def _should_run_pipeline(state: dict[str, Any], run_counter: int, controls: dict[str, Any] | None = None) -> bool:
    """Run once per monotonically increasing sidebar run counter.

    Also guard against accidental duplicate reruns in the same click cycle by
    de-duplicating identical control signatures for the same run counter.
    """
    last_processed = int(state.get("last_processed_run_counter", 0))
    if run_counter <= last_processed:
        return False
    if controls is not None:
        signature = (
            int(run_counter),
            tuple(sorted(str(s) for s in controls.get("sports", []))),
            bool(controls.get("use_ml")),
            bool(controls.get("use_gemini")),
            float(controls.get("bankroll", 0.0)),
            bool(controls.get("theover_spreads") is not None),
            bool(controls.get("theover_totals") is not None),
        )
        if signature == state.get("last_pipeline_signature"):
            logger.info("PIPELINE DIAGNOSTIC: Suppressing duplicate pipeline invocation for identical run signature.")
            return False
        state["last_pipeline_signature"] = signature
    state["last_processed_run_counter"] = run_counter
    return True

def _safe_numeric_series(df: pd.DataFrame, col: str, default: float | int | None = None) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype="float64")
    if col in df.columns:
        s = pd.to_numeric(df[col], errors="coerce")
    else:
        s = pd.Series([pd.NA] * len(df), index=df.index, dtype="Float64")
    if default is not None:
        s = s.fillna(default)
    return s


def _et_floor_day(series: pd.Series) -> pd.Series:
    """Normalize any datetime-like series to ET day boundaries for deterministic joins."""
    return (
        pd.to_datetime(series, errors="coerce", utc=True)
        .dt.tz_convert("America/New_York")
        .dt.floor("D")
    )


def _ml_eligible_market_mask(df: pd.DataFrame) -> pd.Series:
    """Return rows compatible with the shipped ``home_won`` model target."""
    if df is None or df.empty:
        return pd.Series(False, index=getattr(df, "index", pd.RangeIndex(0)), dtype=bool)
    market_type = _safe_str_series(df, "market_type").str.lower()
    return market_type.str.contains(r"moneyline|h2h", regex=True, na=False)


def _compose_model_probability(out: pd.DataFrame) -> tuple[pd.Series, pd.Series, pd.Series]:
    """Expose native ML only for markets compatible with its target.

    Returns a tuple of (model_probability, ml_probability, theover_probability).
    """
    ml = _safe_numeric_series(out, "ml_probability")
    theover = _safe_numeric_series(out, "theover_probability")

    # Apply robust normalization if data format shifts, convert to probability [0, 1]
    # The previous logic only divided by 100 if theover > 1.
    theover = theover.where(theover <= 1.0, theover / 100.0)

    # Reject known broken XGBoost baseline default score when feature matrix collapses.
    is_broken_ml = (ml > 0.19063) & (ml < 0.19064)

    # Log a WARNING with matchup_ids for future retraining tracking
    if is_broken_ml.any():
        broken_matchups = out.loc[is_broken_ml, "matchup_id"].unique() if "matchup_id" in out.columns else []
        logger.warning(f"⚠️ Trapped broken XGBoost scores (0.19063-0.19064) for matchups: {broken_matchups}. Discarding ML predictions and forcing Statistical Fallback.")

    ml_clean = ml.where(~is_broken_ml, pd.NA)
    native_ml = ml_clean.where(_ml_eligible_market_mask(out), pd.NA)
    return native_ml.astype("float64"), native_ml, theover


def _recompute_consensus_from_kalshi(df: pd.DataFrame, require_ml: bool = False) -> pd.DataFrame:
    """Set consensus based on Kalshi availability and probability gap, and update blends."""
    if df is None or df.empty:
        return df
    out = df.copy()

    # Recalculate blended probability and EV/Edge since we might have new Kalshi probabilities
    from core.streamlit_pipeline import compute_blended_probability, _scoped_theover_blend_fade

    kalshi_prob = _safe_numeric_series(out, "kalshi_probability")
    market_prob = _safe_numeric_series(out, "market_probability")

    ml = _safe_numeric_series(out, "ml_probability")
    is_broken_ml = (ml > 0.19063) & (ml < 0.19064)

    if is_broken_ml.any():
        broken_matchups = out.loc[is_broken_ml, "matchup_id"].unique() if "matchup_id" in out.columns else []
        logger.warning(f"⚠️ Consensus Step: Trapped broken XGBoost scores (0.19063-0.19064) for matchups: {broken_matchups}.")

    ml_eligible = _ml_eligible_market_mask(out)
    ml_valid = ml.where(~is_broken_ml & ml_eligible, pd.NA)

    if require_ml and ml_eligible.any() and ml_valid.loc[ml_eligible].notna().sum() == 0:
        raise ValueError("ML predictions failed to merge with the analysis dataframe.")

    # Extract the missing features required for the new Tiered Weight system
    theover_prob = _safe_numeric_series(out, "theover_probability")
    sentiment_prob = _safe_numeric_series(out, "sentiment_diff", default=0.5)

    # FADE genuine-but-cold TheOver sources (model_hit_rate_flipped) in this re-blend,
    # mirroring run_analysis_pipeline. This post-Kalshi refresh recomputes
    # calibrated_probability/EV/blend_in_theover, so it must apply the same fade or it
    # would re-inject TheOver's full Under value and undo the tempering before selection.
    _src_col = out["win_prob_source"] if "win_prob_source" in out.columns else None
    # Mirror run_analysis_pipeline: MLB-tuned fade shrink only on MLB totals (shared helper).
    theover_prob_blend = _scoped_theover_blend_fade(
        theover_prob.to_numpy(dtype=float),
        _src_col,
        _safe_str_series(out, "league"),
        _safe_str_series(out, "market_type"),
        theover_prob.index,
    )

    blended = compute_blended_probability(
        p_market=market_prob,
        p_kalshi=kalshi_prob,
        p_ml=ml_valid,
        p_theover=theover_prob_blend,
        p_sentiment=sentiment_prob,
        league=_safe_str_series(out, "league"),
        market_type=_safe_str_series(out, "market_type")
    )

    # Update core metrics
    out["calibrated_probability"] = blended

    # Refresh blend-input metadata to reflect the final post-Kalshi blend.
    # blend_in_* and blend_tier were recorded in run_analysis_pipeline before
    # live Kalshi was available; now that kalshi_prob is populated, update them
    # so the export/backtest columns accurately describe the actual blend used.
    import numpy as _np
    out["blend_in_kalshi"] = kalshi_prob
    out["blend_in_market"] = market_prob
    out["blend_in_ml"] = ml_valid
    out["blend_in_theover"] = theover_prob_blend
    out["blend_tier"] = _np.where(
        kalshi_prob.fillna(0.0) >= 0.55, 1, 2
    )

    # Check if the Hard Safety Net was used (e.g., probability is exactly 0.5 for all and there's a note)
    # Since ml_valid might be filled with 0.5 from fallback:
    present_ml = _safe_numeric_series(out, "ml_probability").dropna()
    if len(present_ml) > 0 and present_ml.eq(0.5).all():
        logger.warning(
            "Hard Safety Net (Neutral Fallback 0.5) triggered for predictions. "
            "Please check the ML engine logs for the specific missing features that caused the matrix to be mostly empty."
        )

    from core.streamlit_pipeline import american_to_decimal, american_to_prob, get_opposing_odds_from_exchange

    # Always re-derive decimal_odds and market_probability from odds_american to ensure
    # we don't carry stale values after odds patching (e.g. fallback_novig).
    if "odds_american" in out.columns:
        odds_american = _safe_numeric_series(out, "odds_american")
        decimal_odds = odds_american.apply(american_to_decimal)

        # Re-derive market_probability if needed for edge
        implied_prob = odds_american.apply(american_to_prob)
        opposing_implied = odds_american.apply(get_opposing_odds_from_exchange).apply(american_to_prob)
        market_prob = implied_prob / (implied_prob + opposing_implied)
        out["market_probability"] = market_prob
    else:
        decimal_odds = _safe_numeric_series(out, "decimal_odds")

    if decimal_odds.isna().all():
        logger.warning("⚠️ All odds are missing - using default 1.91 (-110 equivalent)")
        decimal_odds = pd.Series([1.91] * len(out), index=out.index)
    # Row-level fallback: ensure every bet row has actionable decimal odds.
    decimal_odds = pd.to_numeric(decimal_odds, errors="coerce").fillna(1.91)

    out["decimal_odds"] = decimal_odds

    # Re-apply the market-anchored MLB total de-bias (#1919/#1921) here too. This
    # post-Kalshi refresh re-blends calibrated_probability and recomputes EV/edge from
    # scratch, so without re-running the correction it silently UNDOES the de-bias that
    # run_analysis_pipeline applied — the card then rebuilds on the raw model over-lean
    # and direction selection never rebalances (the 14 Jun all-Over card: model P(over)
    # ~0.55 vs de-vig market ~0.46 across 15 games, de-bias 0.0747 never reaching the
    # card). Same shared helper as both pipeline paths, so the three cannot drift. Runs
    # before EV/edge/consensus below so all of them see the corrected probability.
    from core.streamlit_pipeline import apply_mlb_total_market_debias
    blended, _mlb_total_debias = apply_mlb_total_market_debias(blended, out)
    out["calibrated_probability"] = blended
    if abs(_mlb_total_debias) > 1e-9:
        out["mlb_total_market_debias"] = _mlb_total_debias

    out["expected_value"] = blended * (decimal_odds - 1) - (1 - blended)
    out["edge"] = blended - market_prob

    status = _safe_str_series(out, "kalshi_match_status").str.lower()

    out["consensus_agreement"] = "No Kalshi"
    valid_kalshi = (kalshi_prob.notna() & (kalshi_prob > 0.0)).fillna(False)
    gap = blended - kalshi_prob

    out.loc[valid_kalshi, "consensus_agreement"] = "Neutral"
    # "Agrees": Kalshi also favors pick direction (P(pick) >= 50%) AND model is more confident.
    # kalshi_probability is pre-oriented to the pick side (P(Under) for Under rows, P(Over)
    # for Over rows) by kalshi_integrator. A value < 0.50 means Kalshi says the OTHER side
    # is more likely — that is Disagrees, not Agrees, regardless of the probability gap.
    out.loc[(valid_kalshi & gap.ge(0.03) & kalshi_prob.ge(0.50)).fillna(False), "consensus_agreement"] = "Agrees"
    # "Disagrees": Kalshi says other direction (P(pick) < 50%) OR Kalshi more confident than model.
    out.loc[(valid_kalshi & (gap.le(-0.03) | kalshi_prob.lt(0.50))).fillna(False), "consensus_agreement"] = "Disagrees"

    # Debug log for probability blend verification (first 5 picks)
    if not out.empty and "market_probability" in out.columns:
        debug_sample = out.head(5)
        for idx, row in debug_sample.iterrows():
            logger.info(
                f"Blend Debug | Pick: {row.get('best_pick', '')} | "
                f"Market: {row.get('market_probability')}, "
                f"Kalshi: {row.get('kalshi_probability')}, "
                f"ML: {row.get('ml_probability')} | "
                f"Blended: {row.get('calibrated_probability')}"
            )

    # Do not filter out rows from the master analysis_df based on edge or expected value.
    # The frontend grids must display the entire master schedule.
    # We leave the strict edge/EV filtering strictly for best_picks_df construction.

    return out

def _merge_kalshi_into_analysis(analysis_df: pd.DataFrame, best_picks_df: pd.DataFrame) -> pd.DataFrame:
    if analysis_df is None or analysis_df.empty or best_picks_df is None or best_picks_df.empty:
        return analysis_df
    kalshi_cols = [
        "kalshi_probability",
        "kalshi_market_title",
        "kalshi_event_ticker",
        "kalshi_match_status",
        "kalshi_match_reason",

    ]
    available_cols = [c for c in kalshi_cols if c in best_picks_df.columns]
    if not available_cols:
        return analysis_df

    merge_keys = ["league", "home_team", "away_team", "game_date"]

    left = analysis_df.copy()
    right = best_picks_df[merge_keys + available_cols].drop_duplicates().copy()

    def _mk_matchup_id(df: pd.DataFrame) -> pd.Series:
        home = df["home_team"].astype(str).str.lower().str.replace(r"[^a-z0-9\s]", "", regex=True).str.strip()
        away = df["away_team"].astype(str).str.lower().str.replace(r"[^a-z0-9\s]", "", regex=True).str.strip()
        team_a = home.where(home <= away, away)
        team_b = away.where(home <= away, home)
        date = pd.to_datetime(df["game_date"], errors="coerce", utc=True)
        date_key = pd.Series([""] * len(df), index=df.index, dtype="string")
        valid = date.notna()
        if valid.any():
            date_key.loc[valid] = date[valid].dt.tz_convert("America/New_York").dt.strftime("%Y-%m-%d")
        return team_a + "|" + team_b + "|" + date_key

    left["matchup_id"] = _mk_matchup_id(left)
    right["matchup_id"] = _mk_matchup_id(right)

    if "game_date" in merge_keys:
        left["game_date"] = pd.to_datetime(left["game_date"], errors="coerce", utc=True).dt.date
        right["game_date"] = pd.to_datetime(right["game_date"], errors="coerce", utc=True).dt.date

    # Create temporary sanitized columns for a bulletproof merge
    left['_merge_home'] = left['home_team'].astype(str).str.lower().str.replace(r'[^a-z0-9]', '', regex=True)
    left['_merge_away'] = left['away_team'].astype(str).str.lower().str.replace(r'[^a-z0-9]', '', regex=True)

    if not right.empty:
        right['_merge_home'] = right['home_team'].astype(str).str.lower().str.replace(r'[^a-z0-9]', '', regex=True)
        right['_merge_away'] = right['away_team'].astype(str).str.lower().str.replace(r'[^a-z0-9]', '', regex=True)

    # Use the temporary sanitized keys instead of the original team names
    sanitized_merge_keys = [k for k in merge_keys if k not in ["home_team", "away_team"]] + ["_merge_home", "_merge_away"]

    merged = left.merge(right, on=sanitized_merge_keys, how="left", suffixes=("", "_best"))

    # Fallback merge path: if strict league-based merge misses, retry using matchup_id (league-agnostic identity).
    if "matchup_id" in left.columns and "matchup_id" in right.columns and merged[available_cols].isna().all(axis=1).any():
        fallback_right = right[["matchup_id", *available_cols]].drop_duplicates("matchup_id")
        fallback_merge = left[["matchup_id"]].merge(fallback_right, on="matchup_id", how="left")
        missing_mask = merged[available_cols].isna().all(axis=1)
        for col in available_cols:
            best_col = f"{col}_best"
            if best_col in merged.columns and col in fallback_merge.columns:
                merged.loc[missing_mask, best_col] = merged.loc[missing_mask, best_col].where(
                    merged.loc[missing_mask, best_col].notna(),
                    fallback_merge.loc[missing_mask, col],
                )

    # Date drift fallback: retry unmatched rows with right-side game_date shifted by +/- 1 day.
    if "game_date" in sanitized_merge_keys and merged[available_cols].isna().all(axis=1).any() and not right.empty:
        shifted_frames = []
        for delta_days in (-1, 1):
            shifted = right.copy()
            shifted["game_date"] = shifted["game_date"].apply(
                lambda d: (d + pd.Timedelta(days=delta_days)).date() if pd.notna(d) else d
            )
            shifted_frames.append(
                left.merge(shifted, on=sanitized_merge_keys, how="left", suffixes=("", "_best"))
            )

        missing_mask = merged[available_cols].isna().all(axis=1)
        for col in available_cols:
            best_col = f"{col}_best"
            if best_col not in merged.columns:
                continue
            for shifted_merge in shifted_frames:
                if best_col in shifted_merge.columns:
                    merged.loc[missing_mask, best_col] = merged.loc[missing_mask, best_col].where(
                        merged.loc[missing_mask, best_col].notna(),
                        shifted_merge.loc[missing_mask, best_col],
                    )

    # Clean up temporary columns
    merged = merged.drop(columns=['_merge_home', '_merge_away'])
    for col in available_cols:
        best_col = f"{col}_best"
        if best_col in merged.columns:
            merged[col] = merged[col].where(merged[col].notna(), merged[best_col]) if col in merged.columns else merged[best_col]
            merged = merged.drop(columns=[best_col])
    return merged




def _sync_ml_probabilities(analysis_df: pd.DataFrame, pipeline_best_picks_df: pd.DataFrame) -> pd.DataFrame:
    """Repair missing native ML probabilities with a market-specific key join."""
    if analysis_df is None or analysis_df.empty or pipeline_best_picks_df is None or pipeline_best_picks_df.empty:
        return analysis_df

    required_cols = ["league", "home_team", "away_team", "game_date", "market_type", "ml_probability"]
    if any(c not in pipeline_best_picks_df.columns for c in required_cols):
        return analysis_df

    left = analysis_df.copy()
    if "ml_probability" in left.columns:
        left.loc[~_ml_eligible_market_mask(left), "ml_probability"] = pd.NA
    right = pipeline_best_picks_df[required_cols].copy()
    right["ml_probability"] = pd.to_numeric(right["ml_probability"], errors="coerce")
    right = right[_ml_eligible_market_mask(right) & right["ml_probability"].notna()].drop_duplicates()
    if right.empty:
        return left

    left["game_date"] = _et_floor_day(left["game_date"])
    right["game_date"] = _et_floor_day(right["game_date"])

    left["_merge_home"] = left["home_team"].astype(str).str.lower().str.replace(r"[^a-z0-9]", "", regex=True)
    left["_merge_away"] = left["away_team"].astype(str).str.lower().str.replace(r"[^a-z0-9]", "", regex=True)
    right["_merge_home"] = right["home_team"].astype(str).str.lower().str.replace(r"[^a-z0-9]", "", regex=True)
    right["_merge_away"] = right["away_team"].astype(str).str.lower().str.replace(r"[^a-z0-9]", "", regex=True)
    left["_merge_market"] = left["market_type"].astype(str).str.lower().str.strip()
    right["_merge_market"] = right["market_type"].astype(str).str.lower().str.strip()

    merge_keys = ["league", "game_date", "_merge_home", "_merge_away", "_merge_market"]
    merged = left.merge(
        right[merge_keys + ["ml_probability"]].rename(columns={"ml_probability": "ml_probability_sync"}),
        on=merge_keys,
        how="left",
    )

    if "ml_probability" in merged.columns:
        merged["ml_probability"] = pd.to_numeric(merged["ml_probability"], errors="coerce")
        merged["ml_probability"] = merged["ml_probability"].where(
            merged["ml_probability"].notna(), merged["ml_probability_sync"]
        )
    else:
        merged["ml_probability"] = merged["ml_probability_sync"]

    recovered = int(pd.to_numeric(merged["ml_probability_sync"], errors="coerce").notna().sum())
    if recovered > 0:
        logger.warning("🔧 ML sync: recovered %s ml_probability values via market-specific merge.", recovered)

    eligible_after_sync = _ml_eligible_market_mask(merged)
    merged.loc[~eligible_after_sync, "ml_probability"] = pd.NA
    ml_after_sync = pd.to_numeric(
        merged.loc[eligible_after_sync, "ml_probability"], errors="coerce"
    ).dropna()
    if len(ml_after_sync) > 1 and ml_after_sync.nunique() <= 1:
        raise ValueError(
            "ML predictions are constant after sync; feature matrix likely invalid from schedule join failure."
        )

    merged = merged.drop(
        columns=[
            col
            for col in ["ml_probability_sync", "_merge_home", "_merge_away", "_merge_market"]
            if col in merged.columns
        ]
    )
    return merged


def _attach_kelly_to_best_picks(best_picks_df: pd.DataFrame, portfolio_df: pd.DataFrame | None, diagnostics: dict) -> pd.DataFrame:
    if best_picks_df is None or best_picks_df.empty:
        return best_picks_df
    out = best_picks_df.copy()
    kelly_map = pd.Series(dtype=float)
    detail_cols = pd.DataFrame()
    if portfolio_df is not None and not portfolio_df.empty and "canonical_pick_key" in portfolio_df.columns:
        detail_cols = (
            portfolio_df.dropna(subset=["canonical_pick_key"])
            .drop_duplicates(subset=["canonical_pick_key"], keep="first")
            .set_index("canonical_pick_key")
        )
        if "production_bet_amount" in detail_cols.columns:
            kelly_map = detail_cols["production_bet_amount"]
        for col in [
            "raw_kelly_amount",
            "production_bet_amount",
            "kelly_cap_reason",
            "production_eligible",
            "kelly_uncalibrated_probability",
            "kelly_probability_used",
            "kelly_probability_source",
        ]:
            if col not in out.columns:
                out[col] = pd.NA
    canonical_key = _safe_str_series(out, "canonical_pick_key").str.strip()
    out["Kelly_Bet_Size"] = canonical_key.map(kelly_map).fillna(0.0)
    if portfolio_df is not None and not portfolio_df.empty and "canonical_pick_key" in portfolio_df.columns:
        out["raw_kelly_amount"] = canonical_key.map(detail_cols["raw_kelly_amount"]) if "raw_kelly_amount" in detail_cols.columns else 0.0
        out["production_bet_amount"] = canonical_key.map(detail_cols["production_bet_amount"]) if "production_bet_amount" in detail_cols.columns else 0.0
        out["kelly_cap_reason"] = canonical_key.map(detail_cols["kelly_cap_reason"]).fillna("") if "kelly_cap_reason" in detail_cols.columns else ""
        out["production_eligible"] = canonical_key.map(detail_cols["production_eligible"]).fillna(out.get("production_eligible", False)) if "production_eligible" in detail_cols.columns else out.get("production_eligible", False)
        for col in [
            "kelly_uncalibrated_probability",
            "kelly_probability_used",
            "kelly_probability_source",
        ]:
            out[col] = canonical_key.map(detail_cols[col]) if col in detail_cols.columns else pd.NA
    status = _safe_str_series(out, "Pick_Status").str.strip()
    line_source = _safe_str_series(out, "market_line_source").str.strip().str.lower()
    line_consistent = pd.Series(out.get("line_consistency_flag", True), index=out.index).fillna(True).astype(bool)
    event_ok = pd.Series(out.get("line_event_identity_match_flag", True), index=out.index).fillna(True).astype(bool)
    pick_text = _safe_str_series(out, "best_pick").str.lower()
    prod_eligible_col = pd.Series(out.get("production_eligible", True), index=out.index).fillna(True).astype(bool)
    safe_mask = status.eq("Actionable") & prod_eligible_col & line_source.eq("live") & line_consistent & event_ok & (~pick_text.str.contains("unresolved", na=False))
    # High Variance/Speculative and Below Threshold picks receive reduced Kelly from the
    # portfolio (30% budget share). Preserve their values unless the line is rejected.
    na_kelly_eligible = status.isin(["High Variance/Speculative", "Below Threshold"])
    rejected_line_mask = line_source.eq("rejected_live") | pick_text.str.contains("unresolved", na=False)
    preserve_na_kelly = na_kelly_eligible & ~rejected_line_mask
    out.loc[~safe_mask & ~preserve_na_kelly, "Kelly_Bet_Size"] = 0.0
    out["kelly_zero_reason"] = ""
    out.loc[status.isin(["No Play", "Missing Line"]), "kelly_zero_reason"] = "non_actionable_status"
    out.loc[na_kelly_eligible & rejected_line_mask, "kelly_zero_reason"] = "non_actionable_rejected_line"
    out.loc[status.eq("Actionable") & (~prod_eligible_col), "kelly_zero_reason"] = "production_ineligible"
    out.loc[status.eq("Actionable") & safe_mask & pd.to_numeric(out["Kelly_Bet_Size"], errors="coerce").fillna(0).le(0), "kelly_zero_reason"] = "zero_after_portfolio_caps"
    out["Kelly_Bet_Size"] = pd.to_numeric(out["Kelly_Bet_Size"], errors="coerce").fillna(0.0).round(2)
    zero_mask = out["Kelly_Bet_Size"].le(0)
    blank_reason = _safe_str_series(out, "kelly_zero_reason").str.strip().eq("")
    out.loc[zero_mask & blank_reason & na_kelly_eligible, "kelly_zero_reason"] = "no_portfolio_allocation"
    blank_reason = _safe_str_series(out, "kelly_zero_reason").str.strip().eq("")
    out.loc[zero_mask & blank_reason, "kelly_zero_reason"] = "zero_without_portfolio_allocation"
    if "kelly_cap_reason" in out.columns:
        out.loc[out["Kelly_Bet_Size"].le(0) & out["kelly_cap_reason"].eq(""), "kelly_cap_reason"] = out.loc[out["Kelly_Bet_Size"].le(0), "kelly_zero_reason"]
    diagnostics["kelly_attached_to_best_picks_count"] = int((out["Kelly_Bet_Size"] > 0).sum())
    diagnostics["kelly_missing_key_count"] = int(canonical_key.eq("").sum())
    diagnostics["kelly_join_match_count"] = int(canonical_key.ne("").sum() - canonical_key[canonical_key.ne("")].map(kelly_map).isna().sum()) if len(canonical_key) else 0
    diagnostics["kelly_join_missing_count"] = int(canonical_key[canonical_key.ne("")].map(kelly_map).isna().sum()) if len(canonical_key) else 0
    diagnostics["kelly_positive_non_actionable_count"] = int(((out["Kelly_Bet_Size"] > 0) & (~status.eq("Actionable"))).sum())
    diagnostics["kelly_best_picks_total_amount"] = float(out["Kelly_Bet_Size"].sum())
    diagnostics["kelly_best_picks_max_amount"] = float(out["Kelly_Bet_Size"].max())
    actionable_zero = status.eq("Actionable") & pd.to_numeric(out["Kelly_Bet_Size"], errors="coerce").fillna(0).le(0)
    diagnostics["actionable_rows_with_zero_kelly_count"] = int(actionable_zero.sum())
    clean_mask = status.eq("Actionable") & line_source.eq("live") & line_consistent & event_ok & (~pick_text.str.contains("unresolved", na=False))
    diagnostics["clean_actionable_rows_with_zero_kelly_count"] = int((clean_mask & pd.to_numeric(out["Kelly_Bet_Size"], errors="coerce").fillna(0).le(0)).sum())
    diagnostics["kelly_zero_reason_counts"] = out.loc[pd.to_numeric(out["Kelly_Bet_Size"], errors="coerce").fillna(0).le(0), "kelly_zero_reason"].value_counts().to_dict()
    return out

def _run_pipeline(controls: dict) -> tuple[dict, list[str], list[str]]:
    """Run the full analysis pipeline. Returns (state_updates, warnings, errors).
    Contains NO st.* calls.
    """
    deferred_warnings: list[str] = []
    deferred_errors: list[str] = []

    spreads_df, err = load_theover_csv(controls.get("theover_spreads"))
    if err:
        deferred_warnings.append(err)

    totals_df, err = load_theover_csv(controls.get("theover_totals"))
    if err:
        deferred_warnings.append(err)

    for upload_df in (spreads_df, totals_df):
        if upload_df is None or upload_df.empty:
            continue
        for team_col in ["home_team", "away_team"]:
            if team_col in upload_df.columns:
                upload_df[team_col] = upload_df[team_col].apply(normalize_team)

    analysis_df, pipeline_best_picks_df, diagnostics = run_analysis_pipeline(
        sports=controls["sports"],
        max_rows=10_000,
        use_ml=bool(controls["use_ml"]),
        spreads_df=spreads_df,
        totals_df=totals_df,
    )

    parlay_columns = ["parlay_legs", "combined_probability", "combined_decimal_odds", "parlay_ev", "legs", "risk_tier", "group_id", "best_payout_book", "Conviction_Score", "min_leg_prob", "has_actionable_anchor", "production_safety_mode", "parlay_class", "premium_eligible", "sellable_as_premium", "commercial_warning", "kelly_fraction", "recommended_bet"]
    empty_per_leg = {f"parlays_{lc}_df": pd.DataFrame(columns=parlay_columns) for lc in (2, 3)}

    empty_state: dict = {
        "analysis_df": pd.DataFrame(),
        "parlays_df": pd.DataFrame(),
        "portfolio_df": pd.DataFrame(),
        "odds_df": pd.DataFrame(),
        "theover_df": pd.DataFrame(),
        "kalshi_df": pd.DataFrame(),
        "gemini_df": pd.DataFrame(),
        "simulation_results": {},
        "best_picks_df": pd.DataFrame(),
        "diagnostics": diagnostics,
        "pipeline_status": "idle",
        "pipeline_running": False,
        **empty_per_leg,
    }

    if analysis_df is None or analysis_df.empty:
        deferred_warnings.append("No rows found for the selected sports.")
        return empty_state, deferred_warnings, deferred_errors

    # Kalshi enrichment with hard timeout
    if isinstance(analysis_df, pd.DataFrame) and not analysis_df.empty:
        if "game_date" not in analysis_df.columns or analysis_df["game_date"].isna().all():
            deferred_warnings.append("game_date missing from analysis_df — Kalshi matching skipped.")
        else:
            analysis_df, kalshi_err = _enrich_with_kalshi_safe(analysis_df)
            if kalshi_err:
                deferred_warnings.append(kalshi_err)

    if controls.get("use_ml"):
        try:
            analysis_df = _sync_ml_probabilities(analysis_df, pipeline_best_picks_df)
        except ValueError as exc:
            deferred_errors.append(f"ML Merge Failed: {exc}")
            return empty_state, deferred_warnings, deferred_errors

        ml_eligible = _ml_eligible_market_mask(analysis_df)
        ml_required = bool(ml_eligible.any())
        eligible_ml_non_null = int(
            _safe_numeric_series(analysis_df, "ml_probability")
            .loc[ml_eligible]
            .notna()
            .sum()
        )
        diagnostics["ml_eligible_rows"] = int(ml_eligible.sum())
        diagnostics["ml_eligible_predictions"] = eligible_ml_non_null
        if ml_required and eligible_ml_non_null == 0:
            deferred_errors.append(
                "ML Merge Failed: Moneyline/H2H predictions could not be joined to the market odds."
            )
            return empty_state, deferred_warnings, deferred_errors
        if not ml_required:
            deferred_warnings.append(
                "ML is enabled, but this slate has no moneyline/H2H rows. "
                "Totals and spreads are continuing with market, Kalshi, and TheOver signals."
            )
    else:
        ml_required = False

    try:
        analysis_df = _recompute_consensus_from_kalshi(
            analysis_df,
            require_ml=ml_required,
        )
    except ValueError as exc:
        deferred_errors.append(f"ML Merge Failed: {exc}")
        return empty_state, deferred_warnings, deferred_errors

    # -----------------------------
    # Update Kalshi Diagnostics
    # -----------------------------
    if "kalshi_probability" in analysis_df.columns:
        if "kalshi_match_status" in analysis_df.columns:
            matched = int(analysis_df["kalshi_match_status"].astype(str).str.lower().eq("matched").sum())
        else:
            matched = int(analysis_df["kalshi_probability"].notna().sum())

        diagnostics["kalshi_matches"] = matched
        diagnostics["kalshi_match_rate"] = matched / max(len(analysis_df), 1)
        if "kalshi_line_diff" in analysis_df.columns and matched > 0:
            avg_diff = analysis_df.loc[analysis_df["kalshi_match_status"].astype(str).str.lower().eq("matched"), "kalshi_line_diff"].mean() if "kalshi_match_status" in analysis_df.columns else analysis_df.loc[analysis_df["kalshi_probability"].notna(), "kalshi_line_diff"].mean()
            diagnostics["kalshi_avg_line_diff"] = avg_diff
            logger.info(f"Kalshi matched {matched} rows with an average line delta of {avg_diff:.4f}")
        else:
            diagnostics["kalshi_avg_line_diff"] = 0.0
    # -----------------------------


    from core.streamlit_pipeline import build_best_picks_df
    from core.streamlit_pipeline import ensure_best_pick_export_columns
    from core.streamlit_pipeline import classify_best_available_picks

    # We pass the diagnostics dictionary so that selection metrics and preview_df
    # can be injected without relying on pandas DataFrame.attrs serialization.
    best_picks_df = build_best_picks_df(analysis_df, diagnostics_out=diagnostics)
    best_picks_df = ensure_best_pick_export_columns(best_picks_df, diagnostics_out=diagnostics)
    diagnostics["identity_columns_ready_before_portfolio"] = bool(
        all(c in best_picks_df.columns for c in ["export_run_id", "pick_id", "canonical_pick_key"])
        and best_picks_df["canonical_pick_key"].astype(str).str.strip().ne("").all()
    )

    if "gemini_analysis" not in analysis_df.columns:
        analysis_df["gemini_analysis"] = ""
    if "gemini_pick" not in analysis_df.columns:
        analysis_df["gemini_pick"] = ""

    # Gemini Integration for Top Picks
    if controls.get("use_gemini") and not best_picks_df.empty:
        try:
            logger.info(f"Firing Gemini API for {len(best_picks_df)} best picks...")
            from integrations.gemini_client import run_gemini_analysis

            # Pass to Gemini wrapper with date columns automatically scrubbed.
            # analysis_df carries every candidate side (not just the winner), so
            # each pick can be paired with its opposing side for a genuine
            # head-to-head comparison instead of a one-sided audit.
            best_picks_df = run_gemini_analysis(best_picks_df, st.session_state, analysis_df=analysis_df)
            logger.info("Gemini analysis payload unpacked successfully.")

            bearish_keywords = ["slow pace", "defensive struggle", "risk of blowout", "key player absences", "grind-it-out"]

            for idx, row in best_picks_df.iterrows():
                risks = str(row.get("gemini_risk_notes", ""))
                pick = str(row.get("gemini_pick", "No Gemini pick"))

                # Phase 4: Qualitative LLM Synergy
                # Apply a 0.85 fractional discount to EV if bearish keywords are detected in the LLM risk notes
                risk_lower = risks.lower()
                if any(kw in risk_lower for kw in bearish_keywords):
                    current_ev = best_picks_df.at[idx, "expected_value"]
                    if pd.notna(current_ev):
                        best_picks_df.at[idx, "expected_value"] = float(current_ev) * 0.85

                # Update analysis_df to reflect these rows were analyzed (for diagnostics tab)
                home = row.get("home_team")
                away = row.get("away_team")
                explanation = row.get("gemini_explanation", "Analyzed")
                if pd.notna(home) and pd.notna(away):
                    mask = (analysis_df["home_team"].eq(home).fillna(False)) & (analysis_df["away_team"].eq(away).fillna(False))
                    analysis_df.loc[mask, "gemini_analysis"] = explanation
                    analysis_df.loc[mask, "gemini_explanation"] = explanation
                    analysis_df.loc[mask, "gemini_risk_notes"] = risks
                    analysis_df.loc[mask, "gemini_pick"] = pick

        except Exception as e:
            deferred_warnings.append(f"Gemini analysis failed: {e}")

    attempted = int(len(analysis_df)) if isinstance(analysis_df, pd.DataFrame) else 0
    if isinstance(analysis_df, pd.DataFrame) and "kalshi_match_status" in analysis_df.columns:
        matched = int(analysis_df["kalshi_match_status"].astype(str).str.lower().eq("matched").sum())
    else:
        matched = int(analysis_df["kalshi_probability"].notna().sum()) if "kalshi_probability" in analysis_df.columns else 0

    diagnostics["kalshi_attempted"] = attempted
    diagnostics["kalshi_matches"] = matched
    diagnostics["kalshi_match_rate"] = float(matched / max(attempted, 1))
    diagnostics["match_rate"] = diagnostics["kalshi_match_rate"]
    diagnostics["kalshi_missing_date_rows"] = int(analysis_df["kalshi_match_reason"].astype(str).eq("missing_date").sum()) if attempted and "kalshi_match_reason" in analysis_df.columns else 0
    diagnostics["kalshi_missing_team_code_rows"] = int(analysis_df["kalshi_match_reason"].astype(str).eq("missing_team_code").sum()) if attempted and "kalshi_match_reason" in analysis_df.columns else 0
    diagnostics["kalshi_no_market_rows"] = int(analysis_df["kalshi_match_reason"].astype(str).eq("no_market_for_tickers").sum()) if attempted and "kalshi_match_reason" in analysis_df.columns else 0
    diagnostics["best_picks"] = int(len(best_picks_df)) if isinstance(best_picks_df, pd.DataFrame) else 0
    diagnostics["positive_ev_rows"] = int((_safe_numeric_series(analysis_df, "expected_value", 0.0) > 0).sum()) if not analysis_df.empty else 0
    diagnostics["positive_ev_picks"] = int((_safe_numeric_series(analysis_df, "expected_value", 0.0) > 0).sum()) if not analysis_df.empty else 0
    diagnostics["best_pick_nonempty_rows"] = int(_safe_str_series(best_picks_df, "best_pick").str.strip().str.len().gt(0).sum()) if not best_picks_df.empty else 0

    parlays_df = generate_parlays(best_picks_df, max_legs=3)
    per_leg: dict = {}
    for lc in (2, 3):
        parlay_slice = parlays_df[_safe_numeric_series(parlays_df, "legs").eq(lc)].copy()
        avail = [c for c in parlay_columns if c in parlay_slice.columns]
        per_leg[f"parlays_{lc}_df"] = parlay_slice[avail] if not parlay_slice.empty else pd.DataFrame(columns=parlay_columns)

    portfolio_df = optimize_portfolio_allocation(best_picks_df, bankroll=float(controls["bankroll"]))
    diagnostics["portfolio_rows_count"] = int(len(portfolio_df)) if isinstance(portfolio_df, pd.DataFrame) else 0
    diagnostics["portfolio_positive_bet_count"] = int((pd.to_numeric(portfolio_df.get("production_bet_amount", 0), errors="coerce").fillna(0).gt(0)).sum()) if isinstance(portfolio_df, pd.DataFrame) and not portfolio_df.empty else 0
    best_picks_df = _attach_kelly_to_best_picks(best_picks_df, portfolio_df, diagnostics)
    best_picks_df = classify_best_available_picks(best_picks_df)
    diagnostics["premium_pick_count"] = int(best_picks_df["sellable_as_premium"].sum())
    diagnostics["best_available_only_count"] = int(best_picks_df["best_available_only"].sum())
    diagnostics["empty_card_recovery_enabled"] = bool(ENABLE_EMPTY_CARD_RECOVERY)
    # Preserve True if the pipeline's internal recovery already fired; only default to False
    diagnostics.setdefault("empty_card_recovery_triggered", False)
    diagnostics["production_card_empty_before_recovery_flag"] = bool(_safe_str_series(best_picks_df, "Pick_Status").eq("Actionable").sum() == 0)
    diagnostics["empty_card_recovery_candidate_count"] = 0
    # Preserve promoted count from pipeline recovery if already set
    diagnostics.setdefault("empty_card_recovery_promoted_count", 0)
    diagnostics["empty_card_recovery_excluded_total_over_count"] = 0
    diagnostics["empty_card_recovery_excluded_line_source_count"] = 0
    diagnostics["empty_card_recovery_excluded_threshold_count"] = 0
    diagnostics["empty_card_recovery_kelly_total"] = 0.0
    diagnostics["production_card_recovery_reason"] = ""

    if ENABLE_EMPTY_CARD_RECOVERY and diagnostics["production_card_empty_before_recovery_flag"]:
        status_s0 = _safe_str_series(best_picks_df, "Pick_Status").str.strip()
        mt0 = _safe_str_series(best_picks_df, "market_type").str.lower()
        src0 = _safe_str_series(best_picks_df, "market_line_source").str.lower()
        lg0 = _safe_str_series(best_picks_df, "league").str.upper()
        pick0 = _safe_str_series(best_picks_df, "best_pick").str.lower()
        line_ok0 = pd.Series(best_picks_df.get("line_consistency_flag", True), index=best_picks_df.index).fillna(True).astype(bool)
        id_ok0 = pd.Series(best_picks_df.get("line_event_identity_match_flag", True), index=best_picks_df.index).fillna(True).astype(bool)
        prod_ev0 = pd.to_numeric(best_picks_df.get("production_expected_value", best_picks_df.get("effective_expected_value", 0)), errors="coerce").fillna(-999)
        prod_edge0 = pd.to_numeric(best_picks_df.get("production_edge", best_picks_df.get("effective_edge", 0)), errors="coerce").fillna(-999)
        prod_prob0 = pd.to_numeric(best_picks_df.get("production_win_probability", best_picks_df.get("effective_win_probability", 0.5)), errors="coerce").fillna(0)
        eff_ev0 = pd.to_numeric(best_picks_df.get("effective_expected_value", best_picks_df.get("expected_value", 0)), errors="coerce").fillna(-999)
        eff_edge0 = pd.to_numeric(best_picks_df.get("effective_edge", best_picks_df.get("edge", 0)), errors="coerce").fillna(-999)
        blocked_stage0 = _safe_str_series(best_picks_df, "status_blocker_stage").str.lower()
        excluded_total_over = mt0.eq("total_over")
        if ALLOW_MLB_TOTAL_OVER_EMPTY_CARD_RECOVERY:
            excluded_total_over = mt0.eq("total_over") & ~lg0.eq("MLB")
        diagnostics["empty_card_recovery_excluded_total_over_count"] = int(excluded_total_over.sum())
        excluded_source = src0.isin([s.lower() for s in EMPTY_CARD_RECOVERY_EXCLUDE_SOURCES]) | ~src0.eq("live")
        diagnostics["empty_card_recovery_excluded_line_source_count"] = int(excluded_source.sum())
        threshold_fail = (prod_ev0 < float(EMPTY_CARD_RECOVERY_MIN_PRODUCTION_EV)) | (prod_edge0 < float(EMPTY_CARD_RECOVERY_MIN_PRODUCTION_EDGE)) | (prod_prob0 < float(EMPTY_CARD_RECOVERY_MIN_PRODUCTION_WIN_PROB))
        diagnostics["empty_card_recovery_excluded_threshold_count"] = int(threshold_fail.sum())
        # Consensus + calibration guards (parity with the build_best_picks_df recovery path):
        # never recover a pick that fades Kalshi (Disagrees / No Kalshi) or whose CALIBRATED
        # win can't beat break-even. Without these this path staked a Disagrees under and a
        # calibrated-negative spread (24 Jun).
        consensus0 = _safe_str_series(best_picks_df, "consensus_agreement").str.strip()
        try:
            from core.streamlit_pipeline import _calibrated_beats_breakeven
            from core.probability_calibration import load_calibration as _load_cal
            from core.empirical_tiers import bucket_key as _bkey, load_bucket_stats as _load_bstats
            from app_core.weights_config import EMPTY_CARD_RECOVERY_CONSENSUS as _RECOVERY_CONSENSUS
            _rec_buckets0 = [
                _bkey(l, m, c) for l, m, c in zip(
                    _safe_str_series(best_picks_df, "league"),
                    _safe_str_series(best_picks_df, "market_type"),
                    _safe_str_series(best_picks_df, "consensus_agreement"),
                )
            ]
            _calib_gate0 = _calibrated_beats_breakeven(
                best_picks_df.get("effective_win_probability", pd.Series(index=best_picks_df.index, dtype=float)),
                best_picks_df.get("odds_american", pd.Series(index=best_picks_df.index, dtype=float)),
                _load_cal(),
                buckets=_rec_buckets0,
                bucket_stats=_load_bstats(),
            ).reindex(best_picks_df.index).fillna(False)
        except Exception as _cal_exc:
            logger.warning("recovery calibration gate unavailable: %s", _cal_exc)
            _calib_gate0 = pd.Series(True, index=best_picks_df.index)
            _RECOVERY_CONSENSUS = ("Agrees", "Neutral")
        consensus_ok0 = consensus0.isin(_RECOVERY_CONSENSUS)
        diagnostics["empty_card_recovery_excluded_calibration_count"] = int(
            (status_s0.isin(["High Variance/Speculative", "Below Threshold"]) & ~_calib_gate0).sum()
        )
        # Owner win-probability floor (parity with the pipeline recovery path,
        # 9 Jul): empirical-first probability must clear MIN_STAKE_WIN_PROBABILITY.
        # Without this, Cincinnati +1.5 (empirical 54.9%, Kalshi 50.5%) re-entered
        # through THIS door at $15 on the strength of the ML blend's 59.8%.
        from app_core.weights_config import MIN_STAKE_WIN_PROBABILITY as _REC_MIN_PROB
        _rec_floor_prob0 = pd.to_numeric(
            best_picks_df.get("empirical_win_probability", pd.Series(index=best_picks_df.index, dtype=float)),
            errors="coerce",
        ).fillna(pd.to_numeric(
            best_picks_df.get("effective_win_probability", pd.Series(index=best_picks_df.index, dtype=float)),
            errors="coerce",
        ))
        recovery_mask = (
            status_s0.isin(["High Variance/Speculative", "Below Threshold"])
            & (~excluded_total_over)
            & (~excluded_source)
            & line_ok0 & id_ok0
            & (~pick0.str.contains("unresolved", na=False))
            & eff_ev0.gt(0) & eff_edge0.gt(0)
            & (~blocked_stage0.isin(["line_provenance", "value_guardrail"]))
            & (~threshold_fail)
            & consensus_ok0
            & _calib_gate0
            & _rec_floor_prob0.ge(float(_REC_MIN_PROB))
        )
        diagnostics["empty_card_recovery_candidate_count"] = int(recovery_mask.sum())
        if recovery_mask.any():
            ranked = best_picks_df[recovery_mask].copy()
            ranked["_rank"] = pd.to_numeric(ranked["Triple_Filter_Rank"], errors="coerce").fillna(9999) if "Triple_Filter_Rank" in ranked.columns else 9999
            ranked["_ev"] = pd.to_numeric(ranked["production_expected_value"], errors="coerce").fillna(-999) if "production_expected_value" in ranked.columns else -999
            ranked["_edge"] = pd.to_numeric(ranked["production_edge"], errors="coerce").fillna(-999) if "production_edge" in ranked.columns else -999
            ranked["_prob"] = pd.to_numeric(ranked["production_win_probability"], errors="coerce").fillna(0) if "production_win_probability" in ranked.columns else 0
            # Win-probability-first (owner doctrine): the recovered pick is the
            # LIKELIEST winner among candidates, not the best-paying one.
            ranked = ranked.sort_values(by=["_prob", "_rank", "_ev", "_edge"], ascending=[False, True, False, False])
            promote_idx = ranked.head(int(EMPTY_CARD_RECOVERY_MAX_PICKS)).index.tolist()
            best_picks_df.loc[promote_idx, "Pick_Status"] = "Actionable"
            best_picks_df.loc[promote_idx, "Status_Reason"] = "Actionable: recovered by empty-card recovery guard"
            best_picks_df.loc[promote_idx, "status_blocker_stage"] = "empty_card_recovery"
            best_picks_df.loc[promote_idx, "status_blocker_reason"] = "Recovered strongest clean non-over candidate after strict guards emptied card"
            best_picks_df.loc[promote_idx, "production_eligible"] = True
            best_picks_df.loc[promote_idx, "kelly_zero_reason"] = ""
            portfolio_df2 = optimize_portfolio_allocation(best_picks_df, bankroll=float(controls["bankroll"]))
            if portfolio_df2 is not None and not portfolio_df2.empty:
                cap_total = float(controls["bankroll"]) * float(EMPTY_CARD_RECOVERY_MAX_KELLY_TOTAL_PCT)
                cap_pick = float(controls["bankroll"]) * float(EMPTY_CARD_RECOVERY_MAX_KELLY_PER_PICK_PCT)
                portfolio_df2["production_bet_amount"] = pd.to_numeric(portfolio_df2.get("production_bet_amount", 0), errors="coerce").fillna(0).clip(upper=cap_pick)
                s = float(portfolio_df2["production_bet_amount"].sum())
                if s > cap_total and s > 0:
                    portfolio_df2["production_bet_amount"] = portfolio_df2["production_bet_amount"] * (cap_total / s)
                portfolio_df = portfolio_df2
                best_picks_df = _attach_kelly_to_best_picks(best_picks_df, portfolio_df, diagnostics)
            diagnostics["empty_card_recovery_triggered"] = True
            diagnostics["empty_card_recovery_promoted_count"] = int(len(promote_idx))
            diagnostics["empty_card_recovery_kelly_total"] = float(pd.to_numeric(best_picks_df.get("Kelly_Bet_Size", 0), errors="coerce").fillna(0).sum())
            diagnostics["production_card_recovery_reason"] = "Recovered strongest clean non-over candidates"
    # Recompute final production-card diagnostics after all guards + Kelly attachment.
    status_s = _safe_str_series(best_picks_df, "Pick_Status").str.strip()
    mt_s = _safe_str_series(best_picks_df, "market_type").str.lower()
    lg_s = _safe_str_series(best_picks_df, "league").str.upper()
    final_actionable_mask = status_s.eq("Actionable")
    final_actionable = best_picks_df[final_actionable_mask]
    final_family_counts = final_actionable["market_type"].astype(str).str.lower().map(
        lambda x: "total" if "total" in x else "side"
    ).value_counts().to_dict() if not final_actionable.empty else {}
    final_type_counts = final_actionable["market_type"].astype(str).str.lower().value_counts().to_dict() if not final_actionable.empty else {}
    diagnostics["actionable_family_counts"] = final_family_counts
    diagnostics["actionable_market_type_counts"] = final_type_counts
    diagnostics["actionable_total_over_count"] = int(final_type_counts.get("total_over", 0))
    diagnostics["actionable_total_under_count"] = int(final_type_counts.get("total_under", 0))
    diagnostics["actionable_side_count"] = int(sum(v for k, v in final_type_counts.items() if "spread" in str(k) or "h2h" in str(k)))
    diagnostics["actionable_mlb_total_over_count"] = int((final_actionable_mask & mt_s.eq("total_over") & lg_s.eq("MLB")).sum())
    diagnostics["totals_only_actionable_flag"] = bool(len(final_actionable) > 0 and diagnostics["actionable_side_count"] == 0)
    diagnostics["viable_side_candidates_count"] = int((mt_s.str.contains("spread|h2h", regex=True, na=False) & status_s.isin(["Actionable", "High Variance/Speculative"])).sum())
    diagnostics["final_actionable_count"] = int(final_actionable_mask.sum())
    diagnostics["final_positive_kelly_count"] = int(pd.to_numeric(best_picks_df.get("Kelly_Bet_Size", 0), errors="coerce").fillna(0).gt(0).sum())
    diagnostics["production_card_empty_flag"] = bool(diagnostics["final_actionable_count"] == 0)
    diagnostics["production_card_empty_after_recovery_flag"] = bool(diagnostics["final_actionable_count"] == 0)
    diagnostics["production_card_empty_reason"] = "No rows survived final production guards" if diagnostics["production_card_empty_flag"] else ""
    if diagnostics["production_card_empty_flag"] and diagnostics.get("empty_card_recovery_enabled"):
        diagnostics["production_card_empty_reason"] = "All candidate rows downgraded by MLB total-over guard, concentration guard, degraded-feature guard, or threshold filters."
    for col, val in {
        "actionable_family_counts": str(final_family_counts),
        "totals_only_actionable_flag": diagnostics["totals_only_actionable_flag"],
        "final_actionable_count": diagnostics["final_actionable_count"],
        "final_positive_kelly_count": diagnostics["final_positive_kelly_count"],
        "production_card_empty_flag": diagnostics["production_card_empty_flag"],
        "production_card_empty_reason": diagnostics["production_card_empty_reason"],
        "clean_actionable_rows_with_zero_kelly_count": diagnostics.get("clean_actionable_rows_with_zero_kelly_count", 0),
        "empty_card_recovery_enabled": diagnostics.get("empty_card_recovery_enabled", False),
        "empty_card_recovery_triggered": diagnostics.get("empty_card_recovery_triggered", False),
        "empty_card_recovery_candidate_count": diagnostics.get("empty_card_recovery_candidate_count", 0),
        "empty_card_recovery_promoted_count": diagnostics.get("empty_card_recovery_promoted_count", 0),
        "empty_card_recovery_excluded_total_over_count": diagnostics.get("empty_card_recovery_excluded_total_over_count", 0),
        "empty_card_recovery_excluded_line_source_count": diagnostics.get("empty_card_recovery_excluded_line_source_count", 0),
        "empty_card_recovery_excluded_threshold_count": diagnostics.get("empty_card_recovery_excluded_threshold_count", 0),
        "empty_card_recovery_kelly_total": diagnostics.get("empty_card_recovery_kelly_total", 0.0),
        "production_card_empty_before_recovery_flag": diagnostics.get("production_card_empty_before_recovery_flag", False),
        "production_card_empty_after_recovery_flag": diagnostics.get("production_card_empty_after_recovery_flag", False),
        "production_card_recovery_reason": diagnostics.get("production_card_recovery_reason", ""),
    }.items():
        best_picks_df[col] = val

    # Recovery can change status and funded stake after the portfolio join; refresh
    # the commercial boundary so the export never carries stale Premium labels.
    best_picks_df = classify_best_available_picks(best_picks_df)
    diagnostics["premium_pick_count"] = int(best_picks_df["sellable_as_premium"].sum())
    diagnostics["best_available_only_count"] = int(best_picks_df["best_available_only"].sum())

    required_portfolio_cols = {"calibrated_probability", "decimal_odds", "recommended_bet"}
    if portfolio_df is not None and not portfolio_df.empty and required_portfolio_cols.issubset(set(portfolio_df.columns)):
        simulation_results = run_bankroll_simulation(portfolio_df, bankroll=float(controls["bankroll"]))
    else:
        diagnostics["bankroll_simulation_skipped"] = True
        simulation_results = {}

    odds_df = analysis_df.copy()
    theover_frames = [spreads_df, totals_df]
    valid_theover_frames = [
        f for f in theover_frames
        if f is not None and isinstance(f, pd.DataFrame) and not f.empty and not f.dropna(how="all").empty
    ]
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        theover_df = pd.concat(valid_theover_frames, ignore_index=True) if valid_theover_frames else pd.DataFrame()

    kalshi_df = analysis_df[analysis_df["kalshi_probability"].notna()].copy() if "kalshi_probability" in analysis_df.columns else analysis_df.iloc[0:0]

    gemini_df = (
        analysis_df[_safe_str_series(analysis_df, "gemini_analysis").str.len() > 0]
        if "gemini_analysis" in analysis_df.columns
        else analysis_df.iloc[0:0]
    )

    # MLB player-prop card (separate from the main best-picks card so a prop-feed hiccup
    # can never break it). Softer market than MLB run totals, so this is where the real
    # edges live; staked at small caps because the prop model is still uncalibrated.
    strikeout_prop_card = pd.DataFrame()
    try:
        from app_core.weights_config import (
            ENABLE_STRIKEOUT_PROPS_PRODUCTION,
            STRIKEOUT_PROP_KELLY_PER_PICK_PCT,
            STRIKEOUT_PROP_KELLY_TOTAL_PCT,
            STRIKEOUT_PROP_KELLY_FRACTION,
        )
        if ENABLE_STRIKEOUT_PROPS_PRODUCTION:
            from datetime import datetime
            import pytz
            from app_core.odds_api import TheOddsAPIClient
            from app_core.prop_runner import build_prop_card
            from core.streamlit_pipeline import _get_odds_api_key

            _prop_key = _get_odds_api_key()
            if _prop_key:
                _prop_date = datetime.now(pytz.timezone("America/New_York")).strftime("%Y-%m-%d")
                strikeout_prop_card = build_prop_card(
                    TheOddsAPIClient(api_key=_prop_key, markets="h2h"),
                    _prop_date,
                    int(_prop_date[:4]),
                    float(controls["bankroll"]),
                    kelly_per_pick_pct=STRIKEOUT_PROP_KELLY_PER_PICK_PCT,
                    kelly_total_pct=STRIKEOUT_PROP_KELLY_TOTAL_PCT,
                    kelly_fraction=STRIKEOUT_PROP_KELLY_FRACTION,
                    prop_results_log=controls.get("prop_results_log"),
                    diagnostics=diagnostics,
                )
                _prop_stake_status = strikeout_prop_card.get(
                    "Stake_Status", pd.Series("", index=strikeout_prop_card.index)
                ).astype(str).str.strip()
                diagnostics["strikeout_prop_actionable_count"] = int(
                    _prop_stake_status.eq("Funded").sum()
                )
                diagnostics["strikeout_prop_research_count"] = int(
                    _prop_stake_status.isin(
                        ["Research / No Stake", "Qualified / No Stake"]
                    ).sum()
                )

                # The legacy strategic-parlay engine only sees game picks. When
                # that board correctly abstains, publish the same strict cross-
                # board duos shown on Best Picks so the parlay export can still
                # use proven player props. Probationary markets are excluded by
                # build_best_duos(strict=True).
                if parlays_df is None or parlays_df.empty:
                    from app_core.best_duos import build_tiered_prop_parlays

                    _prop_parlays = build_tiered_prop_parlays(
                        best_picks_df,
                        strikeout_prop_card,
                        bankroll=float(controls["bankroll"]),
                    )
                    if not _prop_parlays.empty:
                        parlays_df = _prop_parlays
                        diagnostics["strict_prop_parlay_fallback_count"] = int(len(parlays_df))
                        diagnostics["controlled_prop_parlay_count"] = int(
                            _prop_parlays["risk_tier"].eq("Controlled").sum()
                        )
                        diagnostics["research_prop_parlay_count"] = int(
                            _prop_parlays["risk_tier"].eq("Probation / Research").sum()
                        )
                        diagnostics["probation_parlay_fallback"] = bool(
                            _prop_parlays.get(
                                "probation_parlay_mode",
                                pd.Series(False, index=_prop_parlays.index),
                            ).fillna(False).astype(bool).any()
                        )
    except Exception as exc:  # never let the prop slice break the main card
        logger.warning("strikeout prop card build failed: %s", exc)
        diagnostics["strikeout_prop_error"] = str(exc)
        diagnostics["strikeout_prop_feed_status"] = "unexpected_error"
        diagnostics["strikeout_prop_feed_error_type"] = type(exc).__name__
        _prop_error_detail = type(exc).__name__
        if isinstance(exc, SyntaxError):
            _syntax_file = str(getattr(exc, "filename", "") or "").replace("\\", "/").rsplit("/", 1)[-1]
            _syntax_line = getattr(exc, "lineno", None)
            _prop_error_detail = f"SyntaxError in {_syntax_file or 'unknown file'}:{_syntax_line or '?'}"
        diagnostics["strikeout_prop_error_detail"] = _prop_error_detail

    # Final parlay product boundary. Game parlays produced by the strict engine
    # already carry premium_eligible=True. Prop fallback parlays are Premium only
    # when they are Controlled and not probationary; research rows are explicit
    # recreational output and can never carry a funded recommendation.
    if parlays_df is not None and not parlays_df.empty:
        parlays_df = parlays_df.copy()
        probation = pd.Series(
            parlays_df.get("probation_parlay_mode", False), index=parlays_df.index
        ).fillna(False).astype(bool)
        risk_tier = _safe_str_series(parlays_df, "risk_tier").str.strip()
        if "premium_eligible" in parlays_df.columns:
            premium_parlay = pd.Series(
                parlays_df["premium_eligible"], index=parlays_df.index
            ).fillna(False).astype(bool)
        else:
            premium_parlay = risk_tier.eq("Controlled") & ~probation
        parlays_df["premium_eligible"] = premium_parlay
        parlays_df["sellable_as_premium"] = premium_parlay
        parlays_df["parlay_class"] = "Research / Recreational"
        parlays_df.loc[premium_parlay, "parlay_class"] = "Premium"
        if "commercial_warning" not in parlays_df.columns:
            parlays_df["commercial_warning"] = ""
        parlays_df.loc[~premium_parlay, "commercial_warning"] = (
            "Not production-qualified; research/recreational only."
        )
        for stake_col in ("recommended_bet", "kelly_fraction"):
            if stake_col in parlays_df.columns:
                parlays_df.loc[~premium_parlay, stake_col] = 0.0

    # Rebuild per-leg state from the final classified frame, including prop fallback.
    per_leg = {}
    for lc in (2, 3):
        parlay_slice = (
            parlays_df[_safe_numeric_series(parlays_df, "legs").eq(lc)].copy()
            if parlays_df is not None and not parlays_df.empty
            else pd.DataFrame()
        )
        avail = [col for col in parlay_columns if col in parlay_slice.columns]
        per_leg[f"parlays_{lc}_df"] = (
            parlay_slice[avail] if not parlay_slice.empty
            else pd.DataFrame(columns=parlay_columns)
        )
    diagnostics["premium_parlay_count"] = int(
        parlays_df["premium_eligible"].sum()
    ) if parlays_df is not None and not parlays_df.empty else 0
    diagnostics["research_recreational_parlay_count"] = int(
        (~parlays_df["premium_eligible"]).sum()
    ) if parlays_df is not None and not parlays_df.empty else 0

    state_updates = {
        "pipeline_status": "using stored results",
        "pipeline_running": False,
        "strikeout_prop_card": strikeout_prop_card,
        "analysis_df": analysis_df,
        "parlays_df": parlays_df,
        "portfolio_df": portfolio_df,
        "odds_df": odds_df,
        "theover_df": theover_df,
        "kalshi_df": kalshi_df,
        "gemini_df": gemini_df,
        "simulation_results": simulation_results,
        "diagnostics": diagnostics,
        "best_picks_df": best_picks_df,
        **per_leg,
    }
    return state_updates, deferred_warnings, deferred_errors


def main() -> None:
    from app.ui.analysis_dashboard import render_analysis
    try:
        from app.ui.data_diagnostics import show_data_diagnostics
    except Exception:  # pragma: no cover
        def show_data_diagnostics(**_: Any) -> None:
            st.info("Data diagnostics module unavailable in this environment.")
    from app.ui.debug_panel import render_debug, render_debug_panel
    from app.ui.kalshi_diagnostics import render_kalshi_diagnostics
    from app.ui.layout import setup_page
    from app.ui.odds_dashboard import render_odds_table
    from app.ui.sidebar_controls import render_sidebar
    from app.ui.strategy_lab_dashboard import render_strategy_lab

    setup_page()

    stable_defaults = {
        "analysis_df": pd.DataFrame(),
        "best_picks_df": pd.DataFrame(),
        "diagnostics": {},
        "parlays_df": pd.DataFrame(),
        "portfolio_df": pd.DataFrame(),
        "odds_df": pd.DataFrame(),
        "theover_df": pd.DataFrame(),
        "kalshi_df": pd.DataFrame(),
        "gemini_df": pd.DataFrame(),
        "simulation_results": {},
        "pipeline_status": "idle",
        "pipeline_running": False,
    }
    for key, default in stable_defaults.items():
        st.session_state.setdefault(key, default)
    for leg_count in (2, 3):
        st.session_state.setdefault(f"parlays_{leg_count}_df", pd.DataFrame())

    controls = render_sidebar()

    run_counter = int(controls.get("run_analysis_counter", 0))
    should_run = _should_run_pipeline(st.session_state, run_counter, controls)

    # Only run pipeline once per button click; always reset flag on completion or crash
    if should_run and not st.session_state.get("pipeline_running", False):
        st.session_state["pipeline_running"] = True
        try:
            with st.spinner("Running analysis..."):
                state_updates, pipe_warnings, pipe_errors = _run_pipeline(controls)
            st.session_state.update(state_updates)
            for msg in pipe_warnings:
                st.warning(msg)
            for msg in pipe_errors:
                st.error(msg)
        except OddsAPIAuthError as e:
            st.session_state["pipeline_running"] = False
            st.error('The Odds API key is invalid, revoked, or missing. Please verify your credentials in Streamlit secrets.')
            st.stop()
        except Exception:
            st.error(f"Pipeline crashed:\n```\n{traceback.format_exc()}\n```")
        finally:
            # CRITICAL: Always release the pipeline lock so future runs are not blocked
            st.session_state["pipeline_running"] = False

    analysis_df = st.session_state["analysis_df"]

    parlays_df = st.session_state["parlays_df"]
    portfolio_df = st.session_state["portfolio_df"]
    odds_df = st.session_state["odds_df"]
    theover_df = st.session_state["theover_df"]
    kalshi_df = st.session_state["kalshi_df"]
    gemini_df = st.session_state["gemini_df"]
    simulation_results = st.session_state["simulation_results"]

    best_picks_df = st.session_state["best_picks_df"]

    diagnostics = st.session_state.get("diagnostics", {})

    pipeline_status = st.session_state.get("pipeline_status", "idle")
    if not st.session_state["analysis_df"].empty:
        st.caption(f"Pipeline status: {pipeline_status}")





    tab1, tab2, tab3, tab_performance, tab4, tab5, tab6, tab7 = st.tabs(["Odds", "Analysis", "Best Picks", "Performance Recap", "Parlays", "Portfolio", "Debug", "Strategy Lab"])

    with tab_performance:
        from app_core.performance_pipeline import run_performance_pipeline
        from app.ui.results_dashboard import render_results_dashboard

        # Load the performance metrics
        if "performance_df" not in st.session_state:
             with st.spinner("Fetching yesterday's results..."):
                  st.session_state["performance_df"] = run_performance_pipeline()

        perf_df = st.session_state.get("performance_df")
        if perf_df is None:
             from datetime import datetime, timedelta
             yesterday = (datetime.now() - timedelta(days=1)).date()
             st.info(f"No export data found for {yesterday.strftime('%Y-%m-%d')}.")

        # We still want to call render_results_dashboard to show the file uploader
        # even if perf_df is None
        render_results_dashboard(perf_df)

    if analysis_df is None or analysis_df.empty:
        st.info("Configure filters in the sidebar and click **Run Master Analysis**.")
        return


    games_count = int(diagnostics.get("total_games", 0))
    bet_rows = int(diagnostics.get("bet_rows", len(analysis_df)))
    best_rows = int(diagnostics.get("best_picks", len(best_picks_df) if isinstance(best_picks_df, pd.DataFrame) else 0))
    kalshi_matches = int(diagnostics.get("kalshi_matches", 0))
    match_rate = float(diagnostics.get("match_rate", diagnostics.get("kalshi_match_rate", kalshi_matches / max(1, best_rows))))
    totals_games = int(diagnostics.get("theover_totals_games", 0))
    spreads_games = int(diagnostics.get("theover_spreads_games", 0))
    date_fill_attempted = int(diagnostics.get("date_fill_total_rows", 0))
    date_fill_filled = int(diagnostics.get("date_fill_success_rows", 0))
    date_fill_rate = float(diagnostics.get("date_fill_success_rate", 0.0))
    positive_ev_rows = int(diagnostics.get("positive_ev_rows", 0))
    consensus_agrees = (
        int(best_picks_df["consensus_agreement"].astype(str).eq("Agrees").sum())
        if isinstance(best_picks_df, pd.DataFrame) and "consensus_agreement" in best_picks_df.columns
        else 0
    )
    odds_base_loaded = bool(diagnostics.get("odds_schedule_loaded", False))

    with st.container():
        m1, m2, m3, m4, m5, m6, m7, m8, m9, m10, m11 = st.columns(11)
        m1.metric("Total games", games_count)
        m2.metric("Bet rows", bet_rows)
        m3.metric("Best picks", best_rows)
        m4.metric("Kalshi matches", kalshi_matches)
        m5.metric("Match rate", f"{match_rate:.0%}")
        m6.metric("TheOver totals games", f"{totals_games}/{games_count}")
        m7.metric("TheOver spreads games", f"{spreads_games}/{games_count}")
        m8.metric("Date fill success", f"{date_fill_filled}/{date_fill_attempted} ({date_fill_rate:.0%})")
        m9.metric("Positive EV rows", positive_ev_rows)
        m10.metric("Consensus ✅", consensus_agrees)

        kalshi_hits = int(analysis_df["kalshi_match_status"].astype(str).str.lower().eq("matched").sum()) if analysis_df is not None and not analysis_df.empty and "kalshi_match_status" in analysis_df.columns else (
            analysis_df["kalshi_probability"].notna().sum() if analysis_df is not None and not analysis_df.empty and "kalshi_probability" in analysis_df.columns else 0
        )
        total_analysis_len = len(analysis_df) if analysis_df is not None and not analysis_df.empty else 1
        m11.metric("Kalshi Matches", f"{kalshi_hits}/{len(analysis_df) if analysis_df is not None else 0} ({kalshi_hits/total_analysis_len*100:.0f}%)")

        st.progress(max(0.0, min(1.0, match_rate)), text=f"Kalshi match rate: {match_rate:.0%}")
        st.caption(f"Merge keys used: {diagnostics.get('merge_keys_used', [])}")
        st.caption(f"Odds/base schedule loaded: {odds_base_loaded}")
        st.caption(f"Stale base schedule: {bool(diagnostics.get('stale_base_schedule', False))}")
        if diagnostics.get("stale_base_schedule") and diagnostics.get("has_normalized_bet_rows", False):
            st.warning("Pipeline warning: stale base schedule relative to uploaded bet rows.")



    with tab1:
        render_odds_table(analysis_df)

    with tab2:
        if analysis_df is not None and not analysis_df.empty:
            render_analysis(analysis_df)
        if analysis_df is not None and not analysis_df.empty:
            analysis_export_df = analysis_df.copy()
            if "best_pick" in analysis_export_df.columns:
                export_priority = [
                    "league", "home_team", "away_team", "game_date", "game_time_est", "matchup",
                    "market", "pick", "pickteam", "line", "winprobability", "theover_probability",
                    "market_type", "spread_line", "total_line",
                    "expected_value", "edge", "calibrated_probability",
                    "kalshi_probability", "kalshi_market_title", "kalshi_market_ticker", "kalshi_event_ticker", "kalshi_line",
                ]
                ordered_cols = [c for c in export_priority if c in analysis_export_df.columns]
                trailing_cols = [c for c in analysis_export_df.columns if c not in ordered_cols]
                analysis_export_df = analysis_export_df[ordered_cols + trailing_cols]
            analysis_csv = analysis_export_df.to_csv(index=False)
            st.download_button(
                "Export Analysis",
                analysis_csv,
                "analysis_export.csv",
                mime="text/csv",
            )

    with tab3:
        st.subheader("Best Picks")

        # 🏆 Best Overall Pick of the Day — funded production tickets only.
        # Research rows keep grading, but can never become a suggested wager.
        try:
            from app_core.pick_of_day import select_pick_of_the_day
            from app_core.weights_config import MIN_STAKE_WIN_PROBABILITY

            _potd = select_pick_of_the_day(
                best_picks_df, st.session_state.get("strikeout_prop_card")
            )
        except Exception as _potd_exc:  # never let the banner break the tab
            _potd = None
            logger.warning("pick-of-the-day selection failed: %s", _potd_exc)
        if _potd is not None:
            _odds = _potd["odds_american"]
            _odds_txt = "" if _odds is None else f" ({'+' if _odds > 0 else ''}{_odds:.0f})"
            _src = "⚾ MLB Player Prop" if _potd["board"] == "prop" else f"🏟️ {_potd['league']} Game"
            st.success(
                f"🏆 **Best Overall Pick of the Day** — {_src}\n\n"
                f"**{_potd['pick']}**{_odds_txt} · {_potd['detail']}\n\n"
                f"Win probability **{_potd['win_probability']:.1%}** · suggested stake **${_potd['stake']:.2f}**"
            )
            if _potd["below_floor"]:
                st.warning(
                    f"⚠️ Nothing on the board cleared the {float(MIN_STAKE_WIN_PROBABILITY):.0%} "
                    "win-probability floor today — this is the likeliest available winner, "
                    "staked at the flat minimum. Bet light or pass."
                )
            _ru = _potd.get("runner_up")
            if _ru is not None:
                st.caption(
                    f"Runner-up: {_ru['pick']} · {_ru['win_probability']:.1%} "
                    f"({'prop' if _ru['board'] == 'prop' else _ru['league'] + ' game'})"
                )
        else:
            st.info(
                "No production Pick of the Day — no funded game or player-prop "
                "ticket cleared every model, calibration, and portfolio guard."
            )

        # 🎫 Best Duos — likeliest 2-leg parlays across games + props, no shared
        # games between legs, no leg reused across duos (owner request, 8 Jul).
        try:
            from app_core.best_duos import build_best_duos

            _duos = build_best_duos(
                best_picks_df,
                st.session_state.get("strikeout_prop_card"),
                strict=True,
            )
        except Exception as _duo_exc:  # additive; never break the tab
            _duos = None
            logger.warning("best duos failed: %s", _duo_exc)
        if _duos is not None and not _duos.empty:
            st.subheader("🎫 Best Duos — 2-Leg Parlays")
            st.caption(
                "Ranked by conservative post-haircut joint win probability (not payout), legs never share a game "
                "(same-game correlation would silently break the math), and no pick is "
                "reused across duos. Both legs clear the 60% production floor individually."
            )
            for _, _d in _duos.iterrows():
                _pay = f" · pays ${_d['payout_per_10']:.2f} per $10" if pd.notna(_d.get("payout_per_10")) else ""
                st.markdown(
                    f"**{_d['combined_probability']:.1%} combined** · EV {_d.get('parlay_ev', 0.0):+.1%} — "
                    f"{_d['leg1']} ({_d['leg1_prob']:.0%}) **+** {_d['leg2']} ({_d['leg2_prob']:.0%}){_pay}"
                )
            # Owner's ticket (8 Jul): one GAME leg + one PROP leg, best combo.
            try:
                _mixed = build_best_duos(
                    best_picks_df, st.session_state.get("strikeout_prop_card"),
                    max_duos=1, require_mixed=True, strict=True,
                )
            except Exception:
                _mixed = None
            if _mixed is not None and not _mixed.empty:
                _m = _mixed.iloc[0]
                _mpay = f" · pays ${_m['payout_per_10']:.2f} per $10" if pd.notna(_m.get("payout_per_10")) else ""
                st.markdown(
                    f"🎯 **Best Game + Prop duo:** {_m['combined_probability']:.1%} combined · EV {_m.get('parlay_ev', 0.0):+.1%} — "
                    f"{_m['leg1']} ({_m['leg1_prob']:.0%}) **+** {_m['leg2']} ({_m['leg2_prob']:.0%}){_mpay}"
                )
            st.download_button(
                "Export Best Duos",
                _duos.to_csv(index=False, encoding="utf-8-sig"),
                "best_duos.csv",
                mime="text/csv",
                key="export_best_duos",
            )

        # Phase: Sweet Spot Filter Implementation
        with st.expander("🎯 Sweet Spot Filter", expanded=False):
            use_sweet_spot = st.checkbox("Enable Sweet Spot Filter", value=False, key="use_sweet_spot")

            ss_col1, ss_col2, ss_col3 = st.columns(3)
            with ss_col1:
                include_fallback = st.checkbox("Include fallback_novig rows", value=False, key="ss_include_fallback")
                ss_consensus = st.selectbox(
                    "Consensus Requirement",
                    options=["All", "Agrees Only", "Agrees or Neutral"],
                    index=0,
                    key="ss_consensus"
                )
                ss_family = st.selectbox(
                    "Market Family",
                    options=["All", "Overs only", "Unders only", "Sides only"],
                    index=0,
                    key="ss_family"
                )
            with ss_col2:
                ss_min_prob = st.number_input("Min Win Probability", value=0.54, step=0.01, format="%.2f", key="ss_min_prob")
                ss_max_prob = st.number_input("Max Win Probability", value=0.85, step=0.01, format="%.2f", key="ss_max_prob")
                ss_min_edge = st.number_input("Min Edge", value=0.03, step=0.01, format="%.2f", key="ss_min_edge")
            with ss_col3:
                ss_max_edge = st.number_input("Max Edge", value=0.30, step=0.01, format="%.2f", key="ss_max_edge")
                ss_min_ev = st.number_input("Min EV", value=0.04, step=0.01, format="%.2f", key="ss_min_ev")
                ss_max_ev = st.number_input("Max EV", value=0.50, step=0.01, format="%.2f", key="ss_max_ev")

            if use_sweet_spot and best_picks_df is not None and not best_picks_df.empty:
                # Apply filters
                sweet_spot_df = best_picks_df.copy()

                # 1. Actionable + High Variance — Below Threshold explicitly failed
                #    the confidence gate and shouldn't surface as a sweet spot pick.
                sweet_spot_df = sweet_spot_df[
                    sweet_spot_df["Pick_Status"].isin(["Actionable", "High Variance/Speculative"])
                ]
                actionable_count = len(sweet_spot_df)

                # 2. Odds Source
                if not include_fallback:
                    sweet_spot_df = sweet_spot_df[sweet_spot_df["odds_source"] != "fallback_novig"]
                source_count = len(sweet_spot_df)

                # 3. Probability Band
                prob_mask = (sweet_spot_df["calibrated_probability"] >= ss_min_prob) & (sweet_spot_df["calibrated_probability"] <= ss_max_prob)
                sweet_spot_df = sweet_spot_df[prob_mask]
                prob_count = len(sweet_spot_df)

                # 4. Edge Band
                edge_mask = (sweet_spot_df["edge"] >= ss_min_edge) & (sweet_spot_df["edge"] <= ss_max_edge)
                sweet_spot_df = sweet_spot_df[edge_mask]
                edge_count = len(sweet_spot_df)

                # 5. EV Band
                ev_mask = (sweet_spot_df["expected_value"] >= ss_min_ev) & (sweet_spot_df["expected_value"] <= ss_max_ev)
                sweet_spot_df = sweet_spot_df[ev_mask]
                ev_count = len(sweet_spot_df)

                # 6. Consensus Filter
                if ss_consensus == "Agrees Only":
                    sweet_spot_df = sweet_spot_df[sweet_spot_df["consensus_agreement"] == "Agrees"]
                elif ss_consensus == "Agrees or Neutral":
                    sweet_spot_df = sweet_spot_df[sweet_spot_df["consensus_agreement"].isin(["Agrees", "Neutral"])]

                # 7. Family Filter
                if not sweet_spot_df.empty:
                    if ss_family == "Overs only":
                        sweet_spot_df = sweet_spot_df[sweet_spot_df["market_type"] == "total_over"]
                    elif ss_family == "Unders only":
                        sweet_spot_df = sweet_spot_df[sweet_spot_df["market_type"] == "total_under"]
                    elif ss_family == "Sides only":
                        sweet_spot_df = sweet_spot_df[sweet_spot_df["market_type"].str.contains("spread|h2h", na=False)]

                final_count = len(sweet_spot_df)

                st.markdown("#### Diagnostics")
                diag_col1, diag_col2, diag_col3, diag_col4 = st.columns(4)

                diag_col1.metric("Total Best Picks", len(best_picks_df))
                diag_col1.metric("Actionable + High Var", actionable_count)
                diag_col1.metric("After Source Filter", source_count)

                diag_col2.metric("After Prob Band", prob_count)
                diag_col2.metric("After Edge Band", edge_count)
                diag_col2.metric("After EV Band", ev_count)

                diag_col3.metric("Final Sweet Spot", final_count)
                avg_ev = sweet_spot_df["expected_value"].mean() if not sweet_spot_df.empty else 0.0
                avg_edge = sweet_spot_df["edge"].mean() if not sweet_spot_df.empty else 0.0
                diag_col3.metric("Avg EV", f"{avg_ev:.3f}")

                diag_col4.metric("Avg Edge", f"{avg_edge:.3f}")
                min_p = sweet_spot_df["calibrated_probability"].min() if not sweet_spot_df.empty else 0.0
                max_p = sweet_spot_df["calibrated_probability"].max() if not sweet_spot_df.empty else 0.0
                diag_col4.metric("Prob Range", f"{min_p:.2f} - {max_p:.2f}")

                if not sweet_spot_df.empty:
                    m_counts = sweet_spot_df["market_type"].value_counts().to_dict()
                    st.write("**Market Types:**", m_counts)

                    # Ensure export logic uses string columns for status and categorical to prevent issues
                    target_export_cols = [
                        "Pick_Status", "Status_Reason", "Triple_Filter_Rank", "Pick_Quality", "parlay_rank", "league", "Home", "Away", "Local Date",
                        "Commence (Local)", "market_type", "candidate_source", "orientation_source", "upload_match_reason", "best_pick", "WinProbability", "expected_value",
                        "edge", "Conviction_Score", "consensus_agreement", "odds_american", "odds_source", "market_probability",
                        "kalshi_probability", "ml_probability", "gemini_pick", "gemini_explanation", "gemini_risk_notes"
                    ]

                    csv_rename_map = {
                        "home_team": "Home",
                        "away_team": "Away",
                        "game_date": "Local Date",
                        "game_time_est": "Commence (Local)",
                        "calibrated_probability": "WinProbability"
                    }
                    export_prep_ss = sweet_spot_df.rename(columns=csv_rename_map)
                    final_export_cols = [c for c in target_export_cols if c in export_prep_ss.columns]
                    ss_export_csv = export_prep_ss[final_export_cols].to_csv(index=False, encoding="utf-8-sig")

                    st.download_button(
                        "Export Sweet Spot Card",
                        ss_export_csv,
                        "sweet_spot_export.csv",
                        mime="text/csv",
                    )

                    st.markdown("#### Sweet Spot Picks")
                    display_ss = sweet_spot_df.copy()
                    display_ss = display_ss.rename(columns={
                        "Triple_Filter_Rank": "Triple Filter Rank",
                        "Pick_Quality": "Pick Quality",
                        "league": "League",
                        "away_team": "Away Team",
                        "home_team": "Home Team",
                        "game_date": "Game Date",
                        "game_time_est": "Game Time (ET)",
                        "best_pick": "Best Pick",
                        "calibrated_probability": "Prob",
                        "expected_value": "EV",
                        "edge": "Edge",
                        "odds_american": "Odds",
                        "odds_source": "Source",
                        "consensus_agreement": "Consensus",
                        "kalshi_match_status": "Kalshi Status",
                        "ml_probability": "ML Prob",
                    })
                    preferred = ["Pick_Status", "Triple Filter Rank", "Pick Quality", "League", "Home Team", "Away Team", "Game Date", "Game Time (ET)", "Best Pick", "Prob", "ML Prob", "Odds", "Source", "EV", "Edge", "Consensus", "Kalshi Status"]
                    ordered = [c for c in preferred if c in display_ss.columns] + [c for c in display_ss.columns if c not in preferred]
                    st.dataframe(display_ss[ordered], width="stretch")

                    # --- Sweet Spot Parlays ---
                    st.markdown("---")
                    st.markdown("#### 🎯 Sweet Spot Parlays")
                    if len(sweet_spot_df) >= 2:
                        try:
                            from core.probability_calibration import load_calibration
                            from core.smart_parlay_engine import generate_smart_parlays
                            ss_parlays = generate_smart_parlays(
                                sweet_spot_df, num_rr_candidates=5, calibration=load_calibration()
                            )
                            if ss_parlays is not None and not ss_parlays.empty:
                                for leg_count in [2, 3]:
                                    leg_df = ss_parlays[ss_parlays["legs"] == leg_count].head(5).copy()
                                    if leg_df.empty:
                                        continue
                                    st.markdown(f"**Top {len(leg_df)} — {leg_count}-Leg Parlays**")
                                    disp_cols = [c for c in ["parlay_legs", "combined_probability", "combined_decimal_odds", "parlay_ev", "min_leg_prob", "has_actionable_anchor", "best_payout_book", "Conviction_Score"] if c in leg_df.columns]
                                    disp = leg_df[disp_cols].copy()
                                    rename_map = {
                                        "parlay_legs": "Parlay",
                                        "combined_probability": "Hit %",
                                        "combined_decimal_odds": "Payout",
                                        "parlay_ev": "EV",
                                        "min_leg_prob": "Weakest Leg",
                                        "has_actionable_anchor": "Has Actionable",
                                        "best_payout_book": "Book",
                                        "Conviction_Score": "Conviction",
                                    }
                                    disp = disp.rename(columns={k: v for k, v in rename_map.items() if k in disp.columns})
                                    if "Hit %" in disp.columns:
                                        disp["Hit %"] = disp["Hit %"].apply(lambda x: f"{x:.1%}" if pd.notna(x) else "")
                                    if "Payout" in disp.columns:
                                        disp["Payout"] = disp["Payout"].apply(lambda x: f"{x:.2f}x" if pd.notna(x) else "")
                                    if "EV" in disp.columns:
                                        disp["EV"] = disp["EV"].apply(lambda x: f"{x:.3f}" if pd.notna(x) else "")
                                    if "Weakest Leg" in disp.columns:
                                        disp["Weakest Leg"] = disp["Weakest Leg"].apply(lambda x: f"{x:.1%}" if pd.notna(x) else "")
                                    if "Conviction" in disp.columns:
                                        disp["Conviction"] = disp["Conviction"].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "")
                                    st.dataframe(disp, use_container_width=True)
                            else:
                                st.info("No +EV parlays found from Sweet Spot picks — try loosening the filters.")
                        except Exception as e:
                            st.warning(f"Could not generate parlays: {e}")
                    else:
                        st.info("Need at least 2 Sweet Spot picks to build parlays.")
                else:
                    st.info("No picks match the Sweet Spot criteria.")

        with st.expander("Pipeline Debug", expanded=False):
            st.json({
                "kalshi_matches": diagnostics.get("kalshi_matches", 0),
                "kalshi_match_rate": f"{diagnostics.get('kalshi_match_rate', 0.0):.1%}",
                "kalshi_avg_line_diff": round(diagnostics.get("kalshi_avg_line_diff", 0.0), 3),
                "market_type_counts": diagnostics.get("market_type_counts", {}),
                "allowed_market_type_rows": diagnostics.get("allowed_market_type_rows", 0),
                "positive_ev_rows": diagnostics.get("positive_ev_rows", 0),
                "best_pick_nonempty_rows": diagnostics.get("best_pick_nonempty_rows", 0),
                "bet_rows": diagnostics.get("bet_rows", 0),
                "nba_stats_fetch_status": diagnostics.get("nba_stats_fetch_status", "not_started"),
                "nba_stats_fetch_retries_used": diagnostics.get("nba_stats_fetch_retries_used", 0),
                "nba_rows_live_stats": diagnostics.get("nba_rows_live_stats", 0),
                "nba_rows_cached_stats": diagnostics.get("nba_rows_cached_stats", 0),
                "nba_rows_fallback_stats": diagnostics.get("nba_rows_fallback_stats", 0),
                "rows_unresolved_team_mapping": diagnostics.get("rows_unresolved_team_mapping", 0),
                "rows_excluded_from_ml_unresolved_stats": diagnostics.get("rows_excluded_from_ml_unresolved_stats", 0),
                "hybrid_fallback_triggered": diagnostics.get("hybrid_fallback_triggered", False),
                "unmatched_live_games": diagnostics.get("unmatched_live_games", []),
                "missing_uploaded_games": diagnostics.get("missing_uploaded_games", []),
            })

            # Show specific mismatch reports if they exist
            unmatched_live = diagnostics.get("unmatched_live_games", [])
            missing_uploads = diagnostics.get("missing_uploaded_games", [])
            if unmatched_live:
                st.warning(f"Found {len(unmatched_live)} live games that couldn't be matched to uploaded data. Expand to see reasons.")
                st.dataframe(pd.DataFrame(unmatched_live))
            if missing_uploads:
                st.error(f"Found {len(missing_uploads)} uploaded games that did not make it into the final live slate.")
                st.write("Missing Upload IDs:")
                st.write(missing_uploads)

            selection_diags = diagnostics.get("selection_diagnostics", {})
            if selection_diags:
                st.markdown("#### Market-Family Selection Diagnostics")
                col1, col2, col3, col4, col5 = st.columns(5)
                col1.metric("Raw Side", selection_diags.get("raw_family_counts", {}).get("side", 0))
                col1.metric("Raw Total", selection_diags.get("raw_family_counts", {}).get("total", 0))

                col2.metric("Finalist Side", selection_diags.get("finalist_family_counts", {}).get("side", 0))
                col2.metric("Finalist Total", selection_diags.get("finalist_family_counts", {}).get("total", 0))

                col3.metric("Winner Side", selection_diags.get("final_family_counts", {}).get("side", 0))
                col3.metric("Winner Total", selection_diags.get("final_family_counts", {}).get("total", 0))

                col4.metric("Actionable Side", selection_diags.get("actionable_family_counts", {}).get("side", 0))
                col4.metric("Actionable Total", selection_diags.get("actionable_family_counts", {}).get("total", 0))

                avg_scores = selection_diags.get("avg_scores", {})
                col5.metric("Avg Score (Side)", f"{avg_scores.get('side', 0.0):.3f}")
                col5.metric("Avg Score (Total)", f"{avg_scores.get('total', 0.0):.3f}")

                st.markdown("#### Market-Type Detail Counts")

                # Show explicit detail counts per user request
                type_counts = diagnostics.get("market_type_counts", {})
                act_type_counts = diagnostics.get("actionable_market_type_counts", {})

                detail_col1, detail_col2 = st.columns(2)
                detail_col1.write("**Total by Family:**")
                detail_col1.write(f"- Overs: {type_counts.get('total_over', 0)}")
                detail_col1.write(f"- Unders: {type_counts.get('total_under', 0)}")
                detail_col1.write(f"- Sides: {sum(v for k, v in type_counts.items() if 'spread' in k or 'h2h' in k)}")

                detail_col2.write("**Actionable by Family:**")
                detail_col2.write(f"- Actionable Overs: {act_type_counts.get('total_over', 0)}")
                detail_col2.write(f"- Actionable Unders: {act_type_counts.get('total_under', 0)}")
                detail_col2.write(f"- Actionable Sides: {sum(v for k, v in act_type_counts.items() if 'spread' in k or 'h2h' in k)}")

                st.markdown("#### New Calibration Diagnostics")
                cal_col1, cal_col2 = st.columns(2)

                cal_col1.write("**Actionable by League:**")
                league_counts = diagnostics.get("actionable_counts_by_league", {})
                for l, count in league_counts.items():
                    cal_col1.write(f"- {l}: {count}")

                cal_col1.write("**Actionable Totals by League:**")
                total_counts = diagnostics.get("actionable_totals_by_league", {})
                for l, count in total_counts.items():
                    cal_col1.write(f"- {l}: {count}")

                cal_col1.write(f"**Rejected Totals (Prob Floor):** {diagnostics.get('totals_rejected_by_new_guardrails', 0)}")

                cal_col2.write(f"**Fallback-Heavy Slate:** {diagnostics.get('is_fallback_heavy', False)}")
                cal_col2.write(f"**EV Dampener Affected Totals:** {diagnostics.get('ev_dampener_impact_count', 0)}")

                cal_col2.write("**Spread Divergence Override:**")
                cal_col2.write(f"- Downgraded: {diagnostics.get('spreads_downgraded_by_divergence', 0)}")
                cal_col2.write(f"- Rescued: {diagnostics.get('spreads_rescued_by_divergence', 0)}")

                cal_col2.write("**New Guardrails:**")
                cal_col2.write(f"- Blocked by total_under threshold: {diagnostics.get('blocked_by_total_under', 0)}")
                cal_col2.write(f"- Blocked by NHL total penalty: {diagnostics.get('blocked_by_nhl_total', 0)}")
                cal_col2.write(f"- Blocked by MLB spread penalty: {diagnostics.get('blocked_by_mlb_spread_penalty', 0)}")
                cal_col2.write(f"- Blocked by MLB over promotion gate: {diagnostics.get('blocked_by_mlb_over_promotion_gate', 0)}")
                cal_col2.write(f"- Demoted by MLB spread finalist penalty: {diagnostics.get('demoted_by_mlb_spread_finalist_score_penalty', 0)}")
                cal_col2.write(f"- Promoted by NBA side bonus: {diagnostics.get('promoted_by_nba_side_bonus', 0)}")
                cal_col2.write(f"- Promoted by NBA over bonus: {diagnostics.get('promoted_by_nba_over_bonus', 0)}")
                cal_col2.write(f"- Blocked by suspicious-data guardrail: {diagnostics.get('blocked_by_suspicious_data', 0)}")
                cal_col2.write(f"- Rows flagged suspicious_data_flag: {diagnostics.get('suspicious_data_flag_rows', 0)}")
                cal_col2.write(f"- Divergence rows preserved: {diagnostics.get('divergence_rows_preserved', 0)}")
                cal_col2.write(f"- Divergence rows blocked by viability floor: {diagnostics.get('divergence_rows_blocked_by_viability_floor', 0)}")
                cal_col2.write(f"- Divergence rows with negative EV: {diagnostics.get('divergence_rows_negative_ev', 0)}")
                cal_col2.write(f"- Divergence rows with negative edge: {diagnostics.get('divergence_rows_negative_edge', 0)}")

                st.markdown("#### League + Market Calibration Diagnostics")
                cal_lm_col1, cal_lm_col2 = st.columns(2)

                cal_lm_col1.write("**Actionable by League + Market:**")
                actionable_lm = diagnostics.get("actionable_counts_by_league_market", {})
                if actionable_lm:
                    for k, v in actionable_lm.items():
                        cal_lm_col1.write(f"- {k}: {v}")
                else:
                    cal_lm_col1.write("- None")

                cal_lm_col2.write("**Below Threshold by League + Market:**")
                below_lm = diagnostics.get("below_threshold_counts_by_league_market", {})
                if below_lm:
                    for k, v in below_lm.items():
                        cal_lm_col2.write(f"- {k}: {v}")
                else:
                    cal_lm_col2.write("- None")

                st.markdown("#### Consensus & Tuning Diagnostics")
                tun_col1, tun_col2 = st.columns(2)

                tun_col1.write("**Actionable by Consensus:**")
                cons_counts = diagnostics.get("actionable_counts_by_consensus", {})
                for c_name, c_count in cons_counts.items():
                    tun_col1.write(f"- {c_name}: {c_count}")
                tun_col1.write(f"**Final Actionable Count:** {diagnostics.get('final_actionable_count', 0)}")

                tun_col2.write("**Downgrades:**")
                tun_col2.write(f"- Failed Neutral Overlay: {diagnostics.get('downgraded_by_neutral', 0)}")
                tun_col2.write(f"- Failed Disagree Overlay: {diagnostics.get('downgraded_by_disagrees', 0)}")
                tun_col2.write(f"- Failed Side Floor: {diagnostics.get('side_floor_failures', 0)}")

                st.markdown("#### Hidden Bad-Row Check (High Variance with EV<=0 or edge<=0)")
                hidden_bad_rows = diagnostics.get("high_variance_non_positive_ev_edge_rows", [])
                if hidden_bad_rows:
                    st.dataframe(pd.DataFrame(hidden_bad_rows), width="stretch")
                else:
                    st.write("- None")


                st.write("**Pick Status Counts:**")
                if best_picks_df is not None and not best_picks_df.empty and "Pick_Status" in best_picks_df.columns:
                    st.write(best_picks_df["Pick_Status"].value_counts().to_dict())

                st.write("**Consensus Agreement Counts:**")
                if best_picks_df is not None and not best_picks_df.empty and "consensus_agreement" in best_picks_df.columns:
                    st.write(best_picks_df["consensus_agreement"].value_counts().to_dict())

                st.write("**Odds Source Counts:**")
                if best_picks_df is not None and not best_picks_df.empty and "odds_source" in best_picks_df.columns:
                    st.write(best_picks_df["odds_source"].value_counts().to_dict())

                preview_df = selection_diags.get("preview_df")
                if preview_df is not None and not preview_df.empty:
                    st.write("Side vs Total Finalist Preview:")
                    st.dataframe(preview_df, width="stretch")

        with st.expander("📊 Performance Tracker", expanded=False):
            st.markdown("Upload one or more performance recap files (CSV or Excel) to track win rates by status tier over time.")
            recap_files = st.file_uploader(
                "Upload recap file(s)",
                type=["csv", "xlsx", "xls"],
                accept_multiple_files=True,
                key="perf_tracker_uploads",
            )
            if recap_files:
                recap_frames = []
                for f in recap_files:
                    try:
                        if f.name.lower().endswith(".csv"):
                            rdf = pd.read_csv(f)
                        else:
                            rdf = pd.read_excel(f, engine="openpyxl")
                        rdf.columns = [c.strip() for c in rdf.columns]
                        recap_frames.append(rdf)
                    except Exception as e:
                        st.warning(f"Could not parse {f.name}: {e}")
                if recap_frames:
                    recap = pd.concat(recap_frames, ignore_index=True)

                    # Normalise column names — handle both recap format and export format
                    col_map = {c: c for c in recap.columns}
                    if "Pick_Status" in recap.columns and "Status" not in recap.columns:
                        recap = recap.rename(columns={"Pick_Status": "Status"})
                    if "W/L" in recap.columns and "Outcome" not in recap.columns:
                        recap = recap.rename(columns={"W/L": "Outcome"})
                    if "best_pick" in recap.columns and "Pick Taken" not in recap.columns:
                        recap = recap.rename(columns={"best_pick": "Pick Taken"})
                    if "league" in recap.columns and "League" not in recap.columns:
                        recap = recap.rename(columns={"league": "League"})

                    required = {"Outcome", "Status"}
                    missing = required - set(recap.columns)
                    if missing:
                        st.error(
                            f"Uploaded file(s) are missing required columns: {', '.join(sorted(missing))}. "
                            f"Please upload the Performance Recap files (not the Best Picks export). "
                            f"Detected columns: {', '.join(recap.columns.tolist()[:10])}"
                        )
                    else:
                        recap["Outcome"] = recap["Outcome"].astype(str).str.strip().str.upper()
                        recap["Status"] = recap["Status"].astype(str).str.strip()
                        # Drop blank/header rows
                        recap = recap[~recap["Status"].isin(["", "NAN", "STATUS", "NONE"])]
                        recap = recap[recap["Outcome"].isin(["WIN", "LOSS", "W", "L"])]
                        # Normalise W→WIN, L→LOSS
                        recap["Outcome"] = recap["Outcome"].replace({"W": "WIN", "L": "LOSS"})
                        recap["Win"] = recap["Outcome"].eq("WIN")

                        st.markdown(f"**{len(recap)} picks across {len(recap_frames)} recap file(s)**")

                        # --- Staked performance by tier (the headline) ---
                        # The all-rows hit rate blends in Below Threshold / No Play picks the
                        # system declined to stake (≈50% by construction), masking the staked
                        # card. Surface the staked tiers first so good staked days aren't read
                        # as mediocre all-rows days.
                        from app_core.strategy_lab_realized import summarize_recap_tiers
                        tiers = summarize_recap_tiers(recap)
                        st.markdown("#### Staked Performance by Tier")
                        _tcols = st.columns(3)
                        for _col, (_, _row) in zip(_tcols, tiers.iterrows()):
                            _col.metric(
                                _row["Tier"],
                                f"{_row['Hit Rate']:.1%}",
                                f"{int(_row['Wins'])}-{int(_row['Losses'])} ({int(_row['Total'])})",
                            )
                        st.caption(
                            "Judge the system by the Actionable / Actionable+HV tiers. "
                            "'All graded rows' includes Below Threshold and No Play picks that "
                            "were not staked, so it trends toward ~50% by construction."
                        )

                        # Win rate by status
                        status_summary = (
                            recap.groupby("Status")["Win"]
                            .agg(Wins="sum", Total="count")
                            .assign(WinPct=lambda x: x["Wins"] / x["Total"])
                            .reset_index()
                            .sort_values("WinPct", ascending=False)
                        )
                        status_summary["Win %"] = status_summary["WinPct"].map(lambda x: f"{x:.1%}")
                        st.markdown("#### Win Rate by Status")
                        st.dataframe(status_summary[["Status", "Wins", "Total", "Win %"]], use_container_width=True)

                        # Win rate by direction (Over / Under / other)
                        if "Pick Taken" in recap.columns:
                            recap["Direction"] = recap["Pick Taken"].astype(str).apply(
                                lambda p: "Over" if "over" in p.lower() else ("Under" if "under" in p.lower() else "Side")
                            )
                            dir_summary = (
                                recap.groupby(["Direction", "League"])["Win"]
                                .agg(Wins="sum", Total="count")
                                .assign(WinPct=lambda x: x["Wins"] / x["Total"])
                                .reset_index()
                                .sort_values(["League", "Direction"])
                            )
                            dir_summary["Win %"] = dir_summary["WinPct"].map(lambda x: f"{x:.1%}")
                            st.markdown("#### Win Rate by Direction & League")
                            st.dataframe(dir_summary[["League", "Direction", "Wins", "Total", "Win %"]], use_container_width=True)
            else:
                st.info("No recap files uploaded yet.")

        display_df = best_picks_df.copy() if best_picks_df is not None else pd.DataFrame(columns=["league", "pick", "edge"])
        if not display_df.empty and "parlay_rank" in display_df.columns:
            display_df["parlay_rank"] = range(1, len(display_df) + 1)
        # Same display hygiene as the CSV export: $0-stake picks read as "No Bet", not a tier.
        display_df = apply_no_bet_pick_quality(display_df)

        # Keep one directional read per game, but independently label whether the
        # offered price clears the absolute production gate. Nonqualified rows
        # remain visible as BEST AVAILABLE - PASS and carry $0.
        if not display_df.empty:
            try:
                from app_core.lean_card import attach_play_stakes as _aps, score_best_picks_rows as _sbr

                _row_unit = float(st.session_state.get("lean_play_unit", 1.0) or 1.0)
                _scored = _aps(_sbr(best_picks_df), unit=_row_unit)
                if not _scored.empty:
                    display_df["Play Tier"] = _scored["Tier"]
                    display_df["Play Stake"] = _scored["Play_Stake"]
                    display_df["Bet Decision"] = _scored["Bet_Decision"]
                    display_df["Calibrated Edge"] = _scored["Absolute_Edge"]
                    display_df["Production Gate Reason"] = _scored["Production_Gate_Reason"]
            except Exception as _tier_exc:  # tiers are additive; never break the card
                logger.warning("main-card play tiers failed: %s", _tier_exc)

        # Relabel statuses for everything the user reads ("Actionable" is unchanged,
        # so the qualified/no-edge split below still keys off it safely).
        display_df = apply_status_display_labels(display_df)

        if display_df.empty:
            st.warning("⚠️ No games found.")
            st.dataframe(display_df, width="stretch")
        else:
            rename_map = {
                "Triple_Filter_Rank": "Triple Filter Rank",
                "Pick_Quality": "Pick Quality",
                "league": "League",
                "away_team": "Away Team",
                "home_team": "Home Team",
                "game_date": "Game Date",
                "game_time_est": "Game Time (ET)",
                "best_pick": "Best Pick",
                "gemini_pick": "Gemini Pick",
                "calibrated_probability": "Prob",
                "expected_value": "EV",
                "edge": "Edge",
                "odds_american": "Odds",
                "odds_source": "Source",
                "consensus_agreement": "Consensus",
                "kalshi_match_status": "Kalshi Status",
                "ml_probability": "ML Prob",
            }
            display_df = display_df.rename(columns=rename_map)
            if "kalshi_probability" in display_df.columns:
                kalshi_display = pd.to_numeric(display_df["kalshi_probability"], errors="coerce")
                display_df["kalshi_probability_display"] = kalshi_display.map(lambda x: "No Kalshi" if pd.isna(x) else f"{x:.4f}")
            preferred = ["Bet Decision", "Play Tier", "Play Stake", "Pick_Status", "Triple Filter Rank", "Pick Quality", "parlay_rank", "League", "Home Team", "Away Team", "Game Date", "Game Time (ET)", "Best Pick", "Gemini Pick", "Prob", "ML Prob", "Odds", "Source", "EV", "Edge", "Calibrated Edge", "Consensus", "Kalshi Status", "kalshi_probability_display", "Production Gate Reason"]
            ordered = [c for c in preferred if c in display_df.columns] + [c for c in display_df.columns if c not in preferred]
            display_df = display_df[ordered]

            # Display-side cleanup (owner request, 4 Jul): one clear split instead
            # of a wall of No Play / Below Threshold / High Variance statuses.
            # Qualified (Actionable) picks get the table; everything else collapses
            # into a single "No Edge" group with a plain-English reason. Exports
            # and grading keep the raw statuses untouched.
            _disp_status = display_df["Pick_Status"].astype(str).str.strip() if "Pick_Status" in display_df.columns else pd.Series("", index=display_df.index)
            _bet_decision = display_df.get(
                "Bet Decision", pd.Series("BEST AVAILABLE - PASS", index=display_df.index)
            ).astype(str)
            qualified_df = display_df[_disp_status.eq("Actionable") & _bet_decision.eq("BET")]
            no_edge_df = display_df.drop(index=qualified_df.index)

            if not qualified_df.empty:
                st.success(f"✅ {len(qualified_df)} qualified game pick(s) today")
                st.dataframe(qualified_df, width="stretch")
            else:
                st.info(
                    f"🚫 **No production-staked game picks today** — the model has no proven edge "
                    f"at current prices, so the Kelly bankroll sits out. Every game below still "
                    f"shows the best available direction, but failed rows are explicit PASSes at $0. Best action: "
                    f"the 🏆 Pick of the Day above and the ⚾ strikeout props below."
                )

            if not no_edge_df.empty:
                _playable = no_edge_df
                _n_lean = int((_playable.get("Play Tier", pd.Series(dtype=str)) == "LEAN").sum())
                with st.expander(
                    f"📋 Best Available — Pass — {len(no_edge_df)} game(s), {_n_lean} LEAN "
                    f"(all failed the production price gate; tap for reasons)",
                    expanded=False,
                ):
                    no_edge_view = no_edge_df.copy()
                    no_edge_view.insert(0, "Why No Production Bet", no_edge_df.apply(_friendly_no_bet_reason, axis=1))
                    compact_cols = [
                        c for c in [
                            "Bet Decision", "Play Tier", "Play Stake", "Best Pick", "League", "Away Team", "Home Team",
                            "Game Time (ET)", "Prob", "Odds", "EV", "Edge", "Calibrated Edge", "Consensus",
                            "Why No Production Bet", "Pick_Status", "Production Gate Reason",
                        ] if c in no_edge_view.columns
                    ]
                    st.dataframe(no_edge_view[compact_cols], width="stretch")
                    st.caption(
                        "These are directional reads, not wagers. A BET requires positive model EV "
                        "and calibrated probability at least two percentage points above the exact "
                        "sportsbook break-even price. Every PASS remains visible at $0."
                    )

            export_prep_df = best_picks_df.copy()
            # Carry the decision and absolute price-gate evidence into the export.
            if "Play Tier" in display_df.columns:
                export_prep_df["Play_Tier"] = display_df["Play Tier"]
                export_prep_df["Play_Stake"] = display_df["Play Stake"]
                export_prep_df["Bet_Decision"] = display_df["Bet Decision"]
                export_prep_df["Absolute_Edge"] = display_df["Calibrated Edge"]
                export_prep_df["Production_Gate_Reason"] = display_df["Production Gate Reason"]

            csv_rename_map = {
                "home_team": "Home",
                "away_team": "Away",
                "game_date": "Local Date",
                "game_time_est": "Commence (Local)",
                "calibrated_probability": "WinProbability"
            }
            export_prep_df = export_prep_df.rename(columns=csv_rename_map)
            export_prep_df = ensure_best_pick_export_columns(export_prep_df)

            target_export_cols = [
                "Bet_Decision", "Play_Tier", "Play_Stake", "Absolute_Edge", "Production_Gate_Reason",
                "Pick_Status", "Status_Reason", "Triple_Filter_Rank", "Pick_Quality", "parlay_rank", "league", "Home", "Away", "Local Date",
                "Commence (Local)", "market_type", "candidate_source", "orientation_source", "upload_match_reason", "best_pick", "Kelly_Bet_Size", "WinProbability", "expected_value",
                "edge", "Conviction_Score", "consensus_agreement", "odds_american", "odds_source", "market_probability",
                "kalshi_probability", "ml_probability", "gemini_pick", "gemini_explanation", "gemini_risk_notes",
                "status_metric_basis", "effective_expected_value", "effective_edge", "effective_win_probability",
                "empirical_win_probability", "empirical_edge", "empirical_bucket",
                "status_blocker_reason", "status_blocker_stage", "nba_stats_fetch_status", "fallback_summary_by_league",
                "run_health_warning", "degraded_feature_subset_flag", "degraded_feature_subset_reason",
                "production_eligible", "production_win_probability", "production_expected_value", "production_edge",
                "raw_kelly_amount", "production_bet_amount", "kelly_cap_reason", "kelly_zero_reason",
                "final_actionable_count", "final_positive_kelly_count", "production_card_empty_flag", "production_card_empty_reason",
                "clean_actionable_rows_with_zero_kelly_count",
                "empty_card_recovery_enabled", "empty_card_recovery_triggered", "empty_card_recovery_candidate_count",
                "empty_card_recovery_promoted_count", "empty_card_recovery_excluded_total_over_count",
                "empty_card_recovery_excluded_line_source_count", "empty_card_recovery_excluded_threshold_count",
                "empty_card_recovery_kelly_total", "production_card_empty_before_recovery_flag",
                "production_card_empty_after_recovery_flag", "production_card_recovery_reason",
            ]
            for col in REQUIRED_BEST_PICK_EXPORT_COLUMNS:
                if col not in target_export_cols:
                    target_export_cols.append(col)

            missing_required_export_cols = [c for c in REQUIRED_BEST_PICK_EXPORT_COLUMNS if c not in export_prep_df.columns]
            if missing_required_export_cols:
                logger.warning("best_pick_export_missing_columns=%s", missing_required_export_cols)
            logger.info("best_pick_export_required_columns_ok=%s", len(missing_required_export_cols) == 0)

            final_export_cols = [c for c in target_export_cols if c in export_prep_df.columns]
            best_picks_export = export_prep_df[final_export_cols].copy()

            # Phase 5: Output Sanitization (Removed edge/EV filtering to allow neutral/negative fallback games)

            # Apply explicit secondary sorts before export as requested
            # We must preserve the primary Pick_Status > Triple_Filter_Rank > EV > Edge order
            status_order = [
                "Actionable",
                "High Variance/Speculative",
                "Below Threshold",
                "Fallback / Low Confidence",
                "No Play",
                "Missing Line"
            ]
            if "Pick_Status" in best_picks_export.columns:
                best_picks_export["Pick_Status"] = pd.Categorical(best_picks_export["Pick_Status"], categories=status_order, ordered=True)

            # Rank within each tier by WIN PROBABILITY first (owner preference, 3 Jul:
            # the best pick is the one most likely to WIN, not the one that pays best).
            # Probability basis: empirical_win_probability (bucket-realized) with
            # effective_win_probability as fallback; empirical edge / EV / edge remain
            # as tiebreakers only.
            best_picks_export["_prob_sort"] = pd.to_numeric(
                best_picks_export.get("empirical_win_probability"), errors="coerce"
            ).fillna(pd.to_numeric(best_picks_export.get("effective_win_probability"), errors="coerce"))
            best_picks_export["_emp_sort"] = pd.to_numeric(best_picks_export.get("empirical_edge"), errors="coerce")
            best_picks_export["_ev_sort"] = pd.to_numeric(best_picks_export.get("expected_value"), errors="coerce")
            best_picks_export["_edge_sort"] = pd.to_numeric(best_picks_export.get("edge"), errors="coerce")

            sort_cols = ["Pick_Status", "_prob_sort", "_emp_sort", "_ev_sort", "_edge_sort"]
            available_sort_cols = [c for c in sort_cols if c in best_picks_export.columns]

            if available_sort_cols:
                asc = [True, False, False, False, False][:len(available_sort_cols)]
                best_picks_export = best_picks_export.sort_values(available_sort_cols, ascending=asc, na_position="last").reset_index(drop=True)

                # Drop temporary sort columns
                best_picks_export = best_picks_export.drop(columns=["_prob_sort", "_emp_sort", "_ev_sort", "_edge_sort"], errors="ignore")

                if "parlay_rank" in best_picks_export.columns:
                    best_picks_export["parlay_rank"] = range(1, len(best_picks_export) + 1)

            # Display hygiene: benched ($0-stake) picks keep their row but lose the
            # "C-Tier (Value)" style label so the card can't read as a stack of ranked
            # bets when only the staked rows carry money. No change to staking.
            best_picks_export = apply_no_bet_pick_quality(best_picks_export)
            # Exported statuses use the user-facing vocabulary (No Edge / Near Miss /
            # Unproven); applied AFTER the categorical status sort above, which keys
            # off the internal names.
            best_picks_export = apply_status_display_labels(best_picks_export)

            if "Home" in best_picks_export.columns and not best_picks_export.empty:
                if not best_picks_export["Home"].notna().all():
                    st.warning("Warning: Some rows in the Best Picks export have a missing 'Home' team.")

            if "WinProbability" in best_picks_export.columns and not best_picks_export.empty:
                null_pct = best_picks_export["WinProbability"].isna().mean()
                if null_pct > 0.10:
                    st.warning(f"Warning: {null_pct:.1%} of Best Picks are missing ML Probability. Check upstream ML data flow.")

            best_picks_csv = best_picks_export.to_csv(index=False, encoding="utf-8-sig")
            st.download_button(
                "Export Best Picks",
                best_picks_csv,
                "best_picks_export.csv",
                mime="text/csv",
            )

            candidate_audit_df = diagnostics.get("candidate_audit_df")
            if isinstance(candidate_audit_df, pd.DataFrame) and not candidate_audit_df.empty:
                verified_count = int(
                    candidate_audit_df.get(
                        "best_available_selected",
                        pd.Series(False, index=candidate_audit_df.index),
                    ).fillna(False).astype(bool).sum()
                )
                game_count = int(candidate_audit_df["matchup_id"].nunique())
                st.caption(
                    f"Selection audit: {verified_count}/{game_count} games have exactly one "
                    "exported rank-1 winner. Download this file to inspect every candidate, "
                    "the runner-up gap, and why each alternative lost."
                )
                st.download_button(
                    "Export Candidate Selection Audit",
                    candidate_audit_df.to_csv(index=False, encoding="utf-8-sig"),
                    "best_picks_candidate_audit.csv",
                    mime="text/csv",
                    key="export_best_picks_candidate_audit",
                )

            # ── All-games lean view: the model's read on EVERY game, tiered honestly ──
            # Re-presents the same card (no new staking) so a bettor who wants the whole
            # board sees the model's side + confidence + a straight risk label per game.
            try:
                from app_core.lean_card import attach_play_stakes, build_all_games_lean_card
                lean_card = build_all_games_lean_card(best_picks_export)
                if not lean_card.empty:
                    counts = lean_card["Tier"].value_counts().to_dict()
                    st.subheader("🎲 All Games — Play Card")
                    st.caption(
                        f"BET {counts.get('BET', 0)} · LEAN {counts.get('LEAN', 0)} · "
                        f"AVOID {counts.get('AVOID', 0)}.  Ranked by Emp_Edge — the bucket-REALIZED "
                        f"edge (calibrated win vs break-even), NOT model EV (which graded out "
                        f"anti-informative). BET = positive model EV plus at least a 2-point "
                        f"calibrated edge over the exact price. LEAN = its "
                        f"calibrated win beats break-even (your call). AVOID = negative-EV, fading "
                        f"Kalshi, or calibrated win below break-even. Calib_Win% = the model's "
                        f"probability after the bucket-conditional calibration correction. Every valid "
                        f"pregame row remains visible as the Best Available pick, but only absolute-gate "
                        f"BET rows receive a stake; every LEAN/AVOID row is an explicit $0 PASS."
                    )
                    # Keep unit control for qualified BET rows only. LEAN/AVOID
                    # rows stay visible but the absolute gate holds them at $0.
                    play_unit = st.number_input(
                        "Production play unit ($)",
                        min_value=1.0, max_value=100.0, value=1.0, step=1.0,
                        key="lean_play_unit",
                    )
                    lean_card = attach_play_stakes(lean_card, unit=float(play_unit))
                    _play_total = float(lean_card["Play_Stake"].sum())
                    _play_count = int(pd.to_numeric(lean_card["Play_Stake"], errors="coerce").fillna(0).gt(0).sum())
                    st.caption(
                        f"Play_Stake puts ${_play_total:.2f} on {_play_count}/{len(lean_card)} "
                        f"absolute-gate BET rows at a ${play_unit:.0f} unit. LEAN and AVOID rows "
                        f"are best-available directions for reference and remain explicit $0 PASSes."
                    )
                    st.dataframe(lean_card, width="stretch")
                    st.download_button(
                        "Export All-Games Play Card",
                        lean_card.to_csv(index=False, encoding="utf-8-sig"),
                        "all_games_play_card.csv",
                        mime="text/csv",
                    )
            except Exception as exc:  # never let the lean view break the main card
                logger.warning("all-games lean card failed: %s", exc)

            # ── MLB player props (separate softer-market card) ──
            prop_card = st.session_state.get("strikeout_prop_card")
            if prop_card is not None and not prop_card.empty:
                from app_core.prop_runner import stamp_prop_export

                # All funded, all-grading, and research prop CSVs inherit this
                # self-identifying build stamp before they are split.
                prop_card = stamp_prop_export(prop_card, PIPELINE_BUILD)
                _prop_status = prop_card.get(
                    "Stake_Status", pd.Series("", index=prop_card.index)
                ).astype(str).str.strip()
                funded_prop_card = prop_card[_prop_status.eq("Funded")].copy()
                research_prop_card = prop_card[~_prop_status.eq("Funded")].copy()

                st.subheader("⚾ MLB Player Props — Production Picks")
                st.caption(
                    "Core props require at least 62% conservative calibrated win probability; Extended props cover "
                    "the 60–62% band at a flat $1. Both tiers require at least 3% expected "
                    "value, a 0.50-stat model advantage, valid odds, and a non-probation "
                    "market with sufficient graded directional history. The card funds at most two props per game and five total. "
                    "Batter total-base Unders remain research-only until recalibrated. "
                    "Parlays use funded Core and Extended rows only."
                )
                if funded_prop_card.empty:
                    st.info("No player props qualify for a production wager today.")
                else:
                    st.dataframe(funded_prop_card, width="stretch")
                    st.download_button(
                        "Export Funded MLB Player Props",
                        funded_prop_card.to_csv(index=False, encoding="utf-8-sig"),
                        "mlb_player_props_export.csv",
                        mime="text/csv",
                    )

                st.download_button(
                    "Export All MLB Player Props for Next-Day Grading",
                    prop_card.to_csv(index=False, encoding="utf-8-sig"),
                    "mlb_player_props_all_export.csv",
                    mime="text/csv",
                    key="download_all_mlb_props_for_grading",
                )

                if not research_prop_card.empty:
                    with st.expander(
                        f"Research-only props — do not bet ({len(research_prop_card)})",
                        expanded=False,
                    ):
                        st.caption(
                            "These rows remain available for grading and calibration. They are "
                            "not funded and cannot enter recommended parlays."
                        )
                        st.dataframe(research_prop_card, width="stretch")
                        st.download_button(
                            "Export Research-Only MLB Props",
                            research_prop_card.to_csv(index=False, encoding="utf-8-sig"),
                            "mlb_player_props_research_export.csv",
                            mime="text/csv",
                        )
            elif st.session_state.get("diagnostics", {}).get("strikeout_prop_error"):
                _prop_err_type = st.session_state.get("diagnostics", {}).get(
                    "strikeout_prop_error_detail",
                    st.session_state.get("diagnostics", {}).get(
                        "strikeout_prop_feed_error_type", "unexpected error"
                    ),
                )
                st.caption(
                    f"⚾ MLB player props unavailable after retry ({_prop_err_type}). "
                    "Run the slate again; the game card remains valid."
                )
            elif st.session_state.get("diagnostics", {}).get("strikeout_prop_feed_status") in {
                "event_list_failed", "prop_fetch_failed"
            }:
                _prop_stage = st.session_state.get("diagnostics", {}).get("strikeout_prop_feed_status")
                st.caption(
                    f"⚾ Pitcher-prop feed did not respond after retry ({_prop_stage}). "
                    "Run the slate again; no stale props were used."
                )
            else:
                st.caption("⚾ No MLB player props cleared the edge bar today.")

            # Compact export (.xlsx): a readable Excel table with only the columns
            # needed to scan a slate left-to-right, matching the Strategy Lab layout.
            # Win Amount is auto-computed from the Kelly stake and odds; W/L is left
            # blank for manual entry; Total Amount is a P&L formula that resolves
            # once W/L is filled in (+Win Amount on a win, -stake on a loss).
            from io import BytesIO
            import openpyxl
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font

            compact_cols = [
                "WinProbability", "expected_value", "edge",
                "Conviction_Score", "market_probability", "kalshi_probability", "ml_probability",
                "effective_expected_value", "effective_edge", "effective_win_probability",
                "consensus_agreement", "Play_Tier", "Play_Stake", "Pick_Status", "Pick_Quality", "league", "Home", "Away",
                "Commence (Local)", "odds_american", "best_pick", "gemini_pick", "Kelly_Bet_Size",
            ]
            available_compact_cols = [c for c in compact_cols if c in best_picks_export.columns]
            compact_export = best_picks_export[available_compact_cols].copy()
            # Surface the Novig price for the pick as "Novig Line" (the line/total itself is
            # already encoded in best_pick). Sits between "Commence (Local)" and "best_pick"
            # per its position in compact_cols above. odds_american is the Novig price only
            # when odds_source is a genuine Novig quote (odds_api / novig); for uploaded or
            # synthetic-fallback (fallback_novig = -110) rows it is NOT a Novig price, so
            # blank those cells rather than mislabel a non-Novig price as Novig.
            if "odds_american" in compact_export.columns:
                compact_export = compact_export.rename(columns={"odds_american": "Novig Line"})
                if "odds_source" in best_picks_export.columns:
                    _is_novig = best_picks_export["odds_source"].astype(str).str.strip().str.lower().isin({"odds_api", "novig"})
                    compact_export.loc[~_is_novig, "Novig Line"] = pd.NA

            # Win Amount and W/L are left blank for manual entry. Total Amount is a
            # per-row P&L formula (filled below) that resolves once they're entered.
            compact_export["Win Amount"] = ""
            compact_export["W/L"] = ""
            compact_export["Total Amount"] = ""

            final_cols = list(compact_export.columns)
            pct_cols = {
                "WinProbability", "expected_value", "edge", "Conviction_Score",
                "market_probability", "kalshi_probability", "ml_probability",
                "effective_expected_value", "effective_edge", "effective_win_probability",
            }
            money_cols = {"Kelly_Bet_Size", "Win Amount", "Total Amount"}
            odds_cols = {"Novig Line"}  # American odds, shown with explicit sign (e.g. +109 / -114)

            def _col_num(x):
                v = pd.to_numeric(x, errors="coerce")
                return float(v) if pd.notna(v) else None

            def _letter(name):
                return get_column_letter(final_cols.index(name) + 1)

            wl_L = _letter("W/L")
            win_L = _letter("Win Amount")
            kelly_L = _letter("Kelly_Bet_Size") if "Kelly_Bet_Size" in final_cols else None

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Best Picks"
            ws.append(final_cols)

            for i, (_, row) in enumerate(compact_export.iterrows()):
                excel_row = i + 2  # row 1 is the header
                values = []
                for col in final_cols:
                    if col == "Total Amount":
                        values.append(
                            f'=IF({wl_L}{excel_row}="W",{win_L}{excel_row},'
                            f'IF({wl_L}{excel_row}="L",-{kelly_L}{excel_row},""))'
                            if kelly_L else None
                        )
                    elif col == "W/L":
                        values.append(None)
                    elif col in pct_cols or col in money_cols or col in odds_cols:
                        values.append(_col_num(row[col]))
                    else:
                        v = row[col]
                        values.append(None if pd.isna(v) else v)
                ws.append(values)

            last_row = len(compact_export) + 1
            money_fmt = "#,##0"
            pnl_fmt = "#,##0;(#,##0)"  # negatives shown in parentheses, e.g. (2,500)
            for idx, col in enumerate(final_cols, start=1):
                letter = get_column_letter(idx)
                if col in pct_cols:
                    fmt = "0.0%"
                elif col == "Total Amount":
                    fmt = pnl_fmt
                elif col in money_cols:
                    fmt = money_fmt
                elif col in odds_cols:
                    fmt = "+0;-0"  # American odds: +109 / -114
                else:
                    fmt = None
                if fmt:
                    for r in range(2, last_row + 1):
                        ws[f"{letter}{r}"].number_format = fmt
                ws.column_dimensions[letter].width = max(12, min(30, len(col) + 2))

            if len(compact_export) > 0:
                ref = f"A1:{get_column_letter(len(final_cols))}{last_row}"
                table = Table(displayName="BestPicks", ref=ref)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
                )
                ws.add_table(table)

                # Summary rows below the table: "Actionable Totals" (Actionable tier
                # only) and "Totals" (all picks). All values are live formulas over the
                # data range, so they update as Win Amount / W/L are filled in. Net P&L
                # counts a win as +Win Amount and anything not yet won as -stake (money
                # at risk), matching the Strategy Lab convention.
                label_col = "best_pick" if "best_pick" in final_cols else final_cols[0]
                status_L = _letter("Pick_Status") if "Pick_Status" in final_cols else None
                stake_rng = f"{kelly_L}2:{kelly_L}{last_row}"
                win_rng = f"{win_L}2:{win_L}{last_row}"
                wl_rng = f"{wl_L}2:{wl_L}{last_row}"

                def _write_summary(excel_row, label, kelly_f, win_f, wl_f, tot_f):
                    c = ws.cell(row=excel_row, column=final_cols.index(label_col) + 1, value=label)
                    c.font = Font(bold=True)
                    kc = ws.cell(row=excel_row, column=final_cols.index("Kelly_Bet_Size") + 1, value=kelly_f)
                    kc.number_format = money_fmt
                    kc.font = Font(bold=True)
                    wc = ws.cell(row=excel_row, column=final_cols.index("Win Amount") + 1, value=win_f)
                    wc.number_format = money_fmt
                    wc.font = Font(bold=True)
                    rc = ws.cell(row=excel_row, column=final_cols.index("W/L") + 1, value=wl_f)
                    rc.font = Font(bold=True)
                    tc = ws.cell(row=excel_row, column=final_cols.index("Total Amount") + 1, value=tot_f)
                    tc.number_format = pnl_fmt
                    tc.font = Font(bold=True)

                if kelly_L and status_L:
                    act = f'{status_L}2:{status_L}{last_row}'
                    _write_summary(
                        last_row + 2,
                        "Actionable Totals",
                        f'=SUMIF({act},"Actionable",{stake_rng})',
                        f'=SUMIF({act},"Actionable",{win_rng})',
                        f'=COUNTIFS({act},"Actionable",{wl_rng},"W")&"-"&COUNTIFS({act},"Actionable",{wl_rng},"L")',
                        f'=SUMIFS({win_rng},{act},"Actionable",{wl_rng},"W")'
                        f'-SUMIF({act},"Actionable",{stake_rng})'
                        f'+SUMIFS({stake_rng},{act},"Actionable",{wl_rng},"W")',
                    )

                # 🏆 Pick of the Day line so the export carries the day's single
                # best cross-board pick (games + props) alongside the table.
                if _potd is not None:
                    _potd_cell = ws.cell(
                        row=last_row + 4 if kelly_L else last_row + 2,
                        column=final_cols.index(label_col) + 1,
                        value=(
                            f"🏆 Pick of the Day: {_potd['pick']} "
                            f"({_potd['win_probability']:.1%} win prob, ${_potd['stake']:.2f}"
                            f"{', BELOW FLOOR — bet light' if _potd['below_floor'] else ''})"
                        ),
                    )
                    _potd_cell.font = Font(bold=True)

                if kelly_L:
                    _write_summary(
                        last_row + 3,
                        "Totals",
                        f'=SUM({stake_rng})',
                        f'=SUM({win_rng})',
                        f'=COUNTIF({wl_rng},"W")&"-"&COUNTIF({wl_rng},"L")',
                        f'=SUMIF({wl_rng},"W",{win_rng})-SUM({stake_rng})+SUMIF({wl_rng},"W",{stake_rng})',
                    )

            buf = BytesIO()
            wb.save(buf)
            st.download_button(
                "Export Best Picks (Compact)",
                buf.getvalue(),
                "best_picks_compact.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="export_best_picks_compact",
            )

    with tab4:
        st.subheader("Best Parlays")
        base_parlays_df = parlays_df if parlays_df is not None and not parlays_df.empty else pd.DataFrame()

        view_mode = st.radio("Parlay View", ["Ranked Parlays", "Top Combinations"], horizontal=True)

        if base_parlays_df.empty:
            st.info("No parlays available for this view yet.")
        elif view_mode == "Ranked Parlays":
            ranked = base_parlays_df.sort_values("parlay_ev", ascending=False).reset_index(drop=True)
            for idx, row in ranked.iterrows():
                tier = row.get("risk_tier", "")
                tier_label = f" — {tier}" if tier else ""
                st.markdown(f"### Parlay #{idx + 1} ({int(row.get('legs', 0))}-Leg{tier_label})")
                st.markdown(f"- **Combined Probability:** {float(row.get('combined_probability', 0.0)):.2%}")
                st.markdown(f"- **Combined Decimal Odds:** {float(row.get('combined_decimal_odds', 0.0)):.3f}")
                st.markdown(f"- **Parlay EV:** {float(row.get('parlay_ev', 0.0)):.3f}")
                kf = row.get("kelly_fraction", 0.0)
                st.markdown(f"- **Kelly Fraction (1/8):** {float(kf) if pd.notna(kf) else 0.0:.2%}")
                legs = [leg.strip() for leg in str(row.get("parlay_legs", "")).split("|") if leg.strip()]
                for leg in legs:
                    st.markdown(f"- {leg}")
                st.divider()
        else:
            top_combo = base_parlays_df.sort_values("parlay_ev", ascending=False).head(10).reset_index(drop=True)
            table_cols = [c for c in ["combined_probability", "combined_decimal_odds", "parlay_ev", "kelly_fraction", "legs", "risk_tier"] if c in top_combo.columns]
            table_df = top_combo[table_cols].copy()
            table_df.insert(0, "Parlay", ["<br>".join([leg.strip() for leg in str(v).split("|") if leg.strip()]) for v in top_combo["parlay_legs"]])
            st.write(table_df.to_html(escape=False, index=False), unsafe_allow_html=True)

        # Rearrange columns for the export
        parlay_export_columns = [
            "parlay_class", "premium_eligible", "sellable_as_premium", "commercial_warning",
            "risk_tier", "group_id", "parlay_legs", "combined_probability", "combined_decimal_odds",
            "parlay_ev", "legs", "combined_market_prob", "ev_boost_pct", "is_high_correlation",
            "production_safety_mode", "best_payout_book", "Conviction_Score", "min_leg_prob",
            "kelly_fraction", "recommended_bet"
        ]
        export_parlays_df = base_parlays_df.copy() if not base_parlays_df.empty else pd.DataFrame(columns=parlay_export_columns)

        for col in parlay_export_columns:
            if col not in export_parlays_df.columns:
                export_parlays_df[col] = pd.NA

        export_parlays_df = export_parlays_df[parlay_export_columns]

        parlay_csv = export_parlays_df.to_csv(index=False, encoding="utf-8-sig")
        st.download_button(
            "Export Strategic Parlays",
            parlay_csv,
            "smart_parlays_export.csv",
            mime="text/csv",
            key="download_parlays_csv",
        )

    with tab5:
        st.subheader("Portfolio Allocation")
        portfolio_display = portfolio_df.copy() if portfolio_df is not None else pd.DataFrame()
        if not portfolio_display.empty:
            if "best_pick" not in portfolio_display.columns:
                portfolio_display["best_pick"] = ""
            portfolio_display["best_pick"] = _safe_str_series(portfolio_display, "best_pick").str.strip()
            if portfolio_display["best_pick"].str.len().eq(0).all():
                st.warning("Portfolio built, but best_pick strings are missing upstream.")
            display_first_columns = [
                "league", "away_team", "home_team", "game_time_est", "best_pick",
                "calibrated_probability", "expected_value", "edge", "recommended_bet",
            ]
            ordered_columns = [c for c in display_first_columns if c in portfolio_display.columns]
            trailing_columns = [c for c in portfolio_display.columns if c not in ordered_columns]
            portfolio_display = portfolio_display[ordered_columns + trailing_columns]
            league_s = _safe_str_series(portfolio_display, "league").str.upper()
            pick_s = _safe_str_series(portfolio_display, "best_pick")
            bet_s = pd.to_numeric(_safe_str_series(portfolio_display, "recommended_bet", "0"), errors="coerce").fillna(0.0)
            portfolio_display["allocation_label"] = (
                league_s + " | " + pick_s + " | $" + bet_s.map(lambda x: f"{x:,.2f}")
            )
        st.dataframe(portfolio_display, width="stretch")
        st.download_button(
            "Export Portfolio",
            portfolio_display.to_csv(index=False),
            "portfolio_export.csv",
            mime="text/csv",
            key="export_portfolio_csv",
        )

    with tab6:
        show_data_diagnostics(
            odds_df=odds_df,
            theover_df=theover_df if theover_df is not None else analysis_df.iloc[0:0],
            kalshi_df=kalshi_df,
            gemini_df=gemini_df,
        )
        odds_matches = len(odds_df) if odds_df is not None else 0
        theover_matches = len(theover_df) if theover_df is not None else 0
        kalshi_matches_tab = len(kalshi_df) if kalshi_df is not None else 0
        total_analysis_rows = len(analysis_df) if analysis_df is not None else 0
        kalshi_non_null_rows = (
            int(analysis_df["kalshi_probability"].notna().sum())
            if analysis_df is not None and "kalshi_probability" in analysis_df.columns
            else 0
        )
        kalshi_matched_rows = (
            int(analysis_df["kalshi_match_status"].astype(str).str.lower().eq("matched").sum())
            if analysis_df is not None and "kalshi_match_status" in analysis_df.columns
            else 0
        )
        kalshi_miss_rows = (
            int(analysis_df["kalshi_match_status"].astype(str).str.lower().eq("no_match").sum())
            if analysis_df is not None and "kalshi_match_status" in analysis_df.columns
            else 0
        )
        st.markdown("### Kalshi Merge Diagnostics")
        st.write("analysis_df total rows:", total_analysis_rows)
        st.write("rows with non-null kalshi_probability:", kalshi_non_null_rows)
        st.write('rows with kalshi_match_status == "matched":', kalshi_matched_rows)
        st.write('rows with kalshi_match_status == "no_match":', kalshi_miss_rows)
        if controls["show_debug"]:
            render_debug(analysis_df, odds_matches, theover_matches, kalshi_matches_tab)
            render_debug_panel(analysis_df, odds_matches, theover_matches, kalshi_matches_tab)
        else:
            st.info("Enable 'Display Debug Information' in the sidebar to inspect debug data.")
        if controls["show_kalshi_diagnostics"]:
            render_kalshi_diagnostics(analysis_df)
            if analysis_df is not None and not analysis_df.empty and "kalshi_match_status" in analysis_df.columns:
                failures_df = analysis_df[
                    analysis_df["kalshi_match_status"].astype(str).str.lower().ne("matched")
                ].copy()
                failure_cols = [
                    "league",
                    "home_team",
                    "away_team",
                    "kalshi_match_status",
                    "kalshi_match_reason",
                ]
                visible_cols = [c for c in failure_cols if c in failures_df.columns]
                with st.expander("Kalshi Match Failures", expanded=False):
                    if failures_df.empty or not visible_cols:
                        st.info("No unmatched Kalshi rows found.")
                    else:
                        st.dataframe(failures_df[visible_cols], width="stretch")

    with tab7:
        render_strategy_lab(
            analysis_df=analysis_df,
            best_picks_df=best_picks_df,
            portfolio_df=portfolio_df if portfolio_df is not None else analysis_df.iloc[0:0],
            parlays_df=parlays_df if parlays_df is not None else analysis_df.iloc[0:0],
            simulation_results=simulation_results or {},
        )


if __name__ == "__main__":
    main()
